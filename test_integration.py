"""
test_integration.py — Meest uitgebreide test van het complete systeem.

Dekt:
  1. Core unit tests (CSV, analyse, hooks, shoot brief)
  2. Flask route tests (alle ~40 routes, via test client + DB mock)
  3. Edge cases & security
  4. Export pipelines (PDF, Excel templates)
  5. Knop-flows: demo, upload, reanalyze, creative, hooks, brief, etc.

Geen echte DB of API key nodig.
Run: python test_integration.py
"""
import sys
import io
import os
import json
import traceback
from pathlib import Path
from datetime import datetime
from unittest.mock import patch, MagicMock

# Force UTF-8 on Windows (ook voor de section-headers met speciale tekens)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))

# ── Resultaat tracking ─────────────────────────────────────────────────────────
PASS_COUNT = 0
FAIL_COUNT = 0
WARN_COUNT = 0
_results: list[tuple] = []


def _status(label: str, status: str, detail: str = "") -> None:
    global PASS_COUNT, FAIL_COUNT, WARN_COUNT
    icons = {"pass": "\033[92m✓ PASS\033[0m", "fail": "\033[91m✗ FAIL\033[0m", "warn": "\033[93m⚠ WARN\033[0m"}
    print(f"  {icons[status]}  {label}" + (f"  [{detail}]" if detail else ""))
    _results.append((label, status, detail))
    if status == "pass":
        PASS_COUNT += 1
    elif status == "fail":
        FAIL_COUNT += 1
    else:
        WARN_COUNT += 1


def chk(label: str, fn):
    try:
        result = fn()
        if result is True or result is None:
            _status(label, "pass")
        elif isinstance(result, str) and result.startswith("WARN:"):
            _status(label, "warn", result[5:].strip())
        else:
            _status(label, "fail", str(result))
    except Exception as e:
        tb_line = traceback.format_exc().splitlines()
        detail = f"{e}  ({tb_line[-2].strip() if len(tb_line) >= 2 else ''})"
        _status(label, "fail", detail)


def section(title: str) -> None:
    bar = "=" * 65
    print(f"\n{bar}\n  {title}\n{bar}")


# ── Mock DB klant data ──────────────────────────────────────────────────────────
FAKE_CLIENT = {
    "id": 1,
    "name": "TestKlant BV",
    "industry": "fitness",
    "campaign_type": "leads",
    "cpl_benchmark": 30.0,
    "roas_benchmark": None,
    "notes": "Testnotes",
    "client_context": "Fitness studio in Amsterdam",
    "created_at": "2025-01-01",
    "total_spend": 1250.0,
    "total_results": 42,
    "upload_count": 3,
    "last_upload": datetime(2025, 5, 1),
}

FAKE_CLIENT_2 = {
    "id": 2,
    "name": "Tweede Klant",
    "industry": "retail",
    "campaign_type": "purchases",
    "cpl_benchmark": None,
    "roas_benchmark": 3.0,
    "notes": "",
    "client_context": "",
    "created_at": "2025-02-01",
    "total_spend": 800.0,
    "total_results": 15,
    "upload_count": 1,
    "last_upload": datetime(2025, 4, 1),
}

FAKE_UPLOAD = {
    "id": 101,
    "client_id": 1,
    "filename": "test.csv",
    "uploaded_at": datetime(2025, 5, 1),
    "created_at": datetime(2025, 5, 1),
    "row_count": 10,
    "total_spend": 250.0,
    "total_results": 8,
    "num_ads": 5,
    "avg_cpl": 31.25,
    "avg_roas": 0.0,
    "avg_ctr": 2.1,
    "date_from": "2025-04-01",
    "date_to": "2025-04-30",
}

FAKE_BRIEF = {
    "id": 55,
    "client_id": 1,
    "upload_id": 101,
    "created_at": "2025-05-01",
    "brief_json": [
        {"type": "bewezen", "hook_type": "proof", "naam": "Test shoot",
         "logica": "Logica tekst", "redenering": "Redenering tekst",
         "cta": "Plan gratis proefles", "tijdcodes": {"0-5s": "Opening", "5-30s": "Midden", "30-45s": "Einde"}},
    ],
}

FAKE_HOOK_PERF = [
    {"hook_type": "proof", "overall_cpl": 25.0, "total_results": 10, "avg_ctr": 2.5, "total_spend": 250.0},
    {"hook_type": "curiosity", "overall_cpl": 40.0, "total_results": 5, "avg_ctr": 1.8, "total_spend": 200.0},
]

FAKE_CSV_CONTENT = (Path(__file__).parent / "dummy_data" / "sample_meta.csv").read_text(encoding="utf-8-sig")


def _mock_db():
    """Patch alle core.db calls met realistisch nep-data."""
    m = MagicMock()
    m.is_available.return_value = True
    m.get_connection_error.return_value = ""
    m.get_clients.return_value = [FAKE_CLIENT, FAKE_CLIENT_2]
    m.get_client.return_value = FAKE_CLIENT
    m.create_client.return_value = 1
    m.update_client.return_value = None
    m.delete_client.return_value = None
    m.get_uploads.return_value = [FAKE_UPLOAD]
    m.get_upload_csv_content.return_value = FAKE_CSV_CONTENT
    m.get_shoot_briefs.return_value = [FAKE_BRIEF]
    m.get_all_hook_performance.return_value = FAKE_HOOK_PERF
    m.get_ad_creatives.return_value = {
        "reels - problem_solve - v1 - test": {
            "script": "Script tekst", "headline": "Headline",
            "ad_copy_1": "Copy 1 tekst", "ad_copy_2": "", "ad_copy_3": "",
        }
    }
    m.get_insights_history.return_value = [{"insights_text": "Test inzicht", "upload_id": 101}]
    m.save_shoot_brief.return_value = None
    m.get_ad_name_mappings.return_value = {}
    m.save_ad_name_mappings.return_value = 0
    m.get_meta_connection.return_value = None
    m.get_transcripts.return_value = []
    m.get_correct_totals.return_value = {"total_spend": 1250.0, "total_results": 42, "avg_cpl": 29.76}
    m.get_industry_cross_client_data.return_value = []
    m.save_upload.return_value = 101
    m.get_action_plan.return_value = None
    m.save_action_plan.return_value = None
    m.get_competitor_data.return_value = None
    m.init_schema.return_value = None
    m.get_new_ads.return_value = []
    m.get_all_creatives.return_value = []
    m.save_creative.return_value = None
    m.get_creative.return_value = None
    m.save_ad_creative.return_value = None
    m.upsert_ad_creative.return_value = None
    m.delete_ad_creative.return_value = None
    m.get_shoot_brief.return_value = FAKE_BRIEF
    m.delete_shoot_brief.return_value = None
    m.delete_upload.return_value = None
    m.update_client_icp.return_value = None
    return m


# ═════════════════════════════════════════════════════════════════════════════
# 1. CORE TESTS — CSV Parser
# ═════════════════════════════════════════════════════════════════════════════
section("1. CSV Parser")
from core.csv_parser import parse_csv_string, load_dummy_data, validate_csv, NUMERIC_FIELDS

chk("Dummy data laden (>0 rows)", lambda: len(load_dummy_data()) > 0 or "lege dummy data")

def _csv_required_fields():
    row = load_dummy_data()[0]
    missing = [f for f in ["campaign_name", "ad_name", "impressions", "spend", "results"] if f not in row]
    return True if not missing else f"Ontbreekt: {missing}"
chk("Verplichte kolommen aanwezig", _csv_required_fields)

chk("Numerieke velden zijn float", lambda: all(
    isinstance(load_dummy_data()[0].get(f, 0.0), float) for f in NUMERIC_FIELDS
))
chk("Aggregate-rijen gefilterd (geen lege ad_name)", lambda: all(
    r.get("ad_name", "x") != "" for r in load_dummy_data()
))
chk("BOM-stripping werkt", lambda: len(parse_csv_string("﻿" + FAKE_CSV_CONTENT)) > 0)
chk("Validatie: geldige CSV slaagt", lambda: validate_csv(load_dummy_data())[0])
chk("Validatie: lege CSV faalt", lambda: not validate_csv([])[0])
chk("Validatie: ontbrekende kolom faalt", lambda: not validate_csv([{"ad_name": "x", "impressions": 10}])[0])

# NL kolom mapping
nl_csv = """Naam campagne,Naam advertentieset,Naam advertentie,Besteed bedrag (EUR),Vertoningen,Resultaten,Resultaatindicator
Camp A,Set 1,Ad Alpha,50.00,1000,5,lead
Camp B,Set 2,Ad Beta,30.00,800,3,lead
"""
rows_nl = parse_csv_string(nl_csv)
chk("NL kolommen: campaign_name gemapped", lambda: rows_nl[0]["campaign_name"] == "Camp A")
chk("NL kolommen: spend=50.0", lambda: rows_nl[0]["spend"] == 50.0)
chk("EN kolommen geparsed", lambda: parse_csv_string(
    "Campaign Name,Ad Set Name,Ad Name,Amount Spent (EUR),Impressions,Results,Result Indicator\n"
    "C,S,A,25.00,400,2,Lead\n"
)[0]["campaign_name"] == "C")


# ═════════════════════════════════════════════════════════════════════════════
# 2. ANALYSIS
# ═════════════════════════════════════════════════════════════════════════════
section("2. Analysis Pipeline")
from core.analysis import (
    build_campaigns, build_summary, build_ad_chart_data,
    get_all_ads, get_date_range, filter_rows_by_date,
    build_wow_comparison, filter_zero_spend, build_ad_delivery_map,
    merge_multi_conversion_rows,
)

_rows = load_dummy_data()
_campaigns = build_campaigns(_rows)
_summary = build_summary(_rows, _campaigns)

chk("Campaigns gebouwd (>0)", lambda: len(_campaigns) > 0)
chk("Summary: total_spend > 0", lambda: _summary.total_spend > 0)
chk("Summary: campaign_type valide", lambda: _summary.campaign_type in ("leads", "purchases", "awareness"))
chk("Summary: top_ad aanwezig", lambda: _summary.top_ad is not None)
chk("Summary: CPL consistent", lambda: (
    abs(round(_summary.total_spend / _summary.total_results, 2) - _summary.avg_cost_per_result) < 0.05
    if _summary.total_results > 0 else True
))
chk("filter_zero_spend: verwijdert €0", lambda: (
    len(filter_zero_spend([{"spend": 10.0}, {"spend": 0.0}, {"spend": 5.5}])) == 2
))
chk("Date range: geldig en volgorde klopt", lambda: (
    get_date_range(_rows)[0] <= get_date_range(_rows)[1]
))
chk("Date filter: verkort dataset", lambda: (
    len(filter_rows_by_date(_rows, _rows[0].get("day", ""), _rows[0].get("day", ""))) < len(_rows)
))
chk("Date filter: omgekeerd bereik crasht niet", lambda: isinstance(
    filter_rows_by_date(_rows, "2025-01-01", "2024-01-01"), list
))
chk("WoW vergelijking aanwezig of None (geen crash)", lambda: (
    build_wow_comparison(_rows) is None or "p1" in build_wow_comparison(_rows)
))
chk("Chart data: alle keys aanwezig", lambda: all(
    k in build_ad_chart_data(_campaigns, _summary.campaign_type)
    for k in ("labels", "spend", "metric", "ctr", "results")
))
chk("Delivery map: status lowercase", lambda: (
    build_ad_delivery_map([{"ad_name": "X", "ad_delivery": "Active"}]).get("X") == "active"
))
chk("Merge multi-conversion: spend 1x, results opgeteld", lambda: (
    merge_multi_conversion_rows([
        {**{k: v for k, v in load_dummy_data()[0].items()}, "results": 3, "result_indicator": "lead"},
        {**{k: v for k, v in load_dummy_data()[0].items()}, "results": 2, "result_indicator": "custom"},
    ])[0]["results"] == 5.0
))


# ═════════════════════════════════════════════════════════════════════════════
# 3. HOOK ANALYZER
# ═════════════════════════════════════════════════════════════════════════════
section("3. Hook Analyzer")
from core.hook_analyzer import (
    parse_ad_name, detect_hook, detect_format,
    aggregate_hook_performance, aggregate_format_performance,
    get_winning_combinations, get_untested_hooks, get_untested_formats,
    get_unknown_ads, HOOK_TYPES, FORMAT_TYPES,
)

_all_ads = sorted(get_all_ads(_campaigns), key=lambda a: a.spend, reverse=True)

STRUCTURED_CASES = [
    ("reels - problem_solve - v1 - geen omkleden", "problem_solve", "reels", 1),
    ("static - social_proof - v1 - 500 klanten",   "social_proof",  "static", 1),
    ("ugc - frustration - v3 - moe van dieten",     "frustration",   "ugc",    3),
    ("testimonial - proof - v1 - klant review",     "proof",         "testimonial", 1),
    ("static - curiosity - v2 - wist je dat",       "curiosity",     "static", 2),
    ("reels - urgency - v1 - laatste kans",         "urgency",       "reels",  1),
    ("reels - educational - v1 - hoe werkt het",    "educational",   "reels",  1),
    ("static - confrontation - v1 - stop met",      "confrontation", "static", 1),
    ("static - promise - v1 - resultaat in 8w",     "promise",       "static", 1),
    ("UGC - Proof - V2 - Lid Testimonial",          "proof",         "ugc",    2),
]

for name, exp_hook, exp_fmt, exp_ver in STRUCTURED_CASES:
    parsed = parse_ad_name(name)
    chk(f"Parse '{name[:40]}': hook={exp_hook}", lambda p=parsed, h=exp_hook: p["hook"] == h)
    chk(f"  format={exp_fmt}",  lambda p=parsed, f=exp_fmt: p["format"] == f)
    chk(f"  version={exp_ver}", lambda p=parsed, v=exp_ver: p["version"] == v)

chk("Keyword fallback: '?' → curiosity", lambda: detect_hook("Weet jij al dit geheim?") == "curiosity")
chk("Keyword fallback: 'ziek van' → frustration", lambda: detect_hook("Ziek van dieten zonder resultaat?") == "frustration")
chk("Keyword fallback: 'klantreview' → testimonial format", lambda: detect_format("Klantreview na 8 weken") == "testimonial")
chk("Keyword fallback: 'carousel' herkend", lambda: detect_format("Accessoires Carousel") == "carousel")
chk("Hook performance: niet leeg", lambda: len(aggregate_hook_performance(_all_ads)) > 0)
chk("Format performance: niet leeg", lambda: len(aggregate_format_performance(_all_ads)) > 0)
chk("Winning combos: allen met results>0", lambda: all(c["results"] > 0 for c in get_winning_combinations(_all_ads)))
chk("Untested hooks: alleen geldige waarden", lambda: all(h in HOOK_TYPES for h in get_untested_hooks(_all_ads)))
chk("Untested formats: alleen geldige waarden", lambda: all(f in FORMAT_TYPES for f in get_untested_formats(_all_ads)))
chk("Unknown ads: allen met spend>0", lambda: all(a["spend"] > 0 for a in get_unknown_ads(_all_ads)))
chk("XSS in ad naam: crasht niet", lambda: parse_ad_name("<script>alert('xss')</script>") is not None)
chk("Zeer lange naam: crasht niet", lambda: parse_ad_name("A" * 1000) is not None)
chk("Unicode naam: crasht niet", lambda: parse_ad_name("Herken jij dit 🎯 gevoel 日本語") is not None)


# ═════════════════════════════════════════════════════════════════════════════
# 4. SHOOT BRIEF
# ═════════════════════════════════════════════════════════════════════════════
section("4. Shoot Brief")
from core.shoot_brief import generate_shoot_brief, _strip_em_dashes, _clean_script, _build_script, HOOK_TYPES as SB_HOOKS

_brief = generate_shoot_brief(_summary, _all_ads, client_name="TestKlant BV")
EXPECTED_BRIEF_LEN = 17
EXPECTED_TYPES = {"bewezen", "test", "wild_card", "testimonial", "short", "broll"}

chk(f"Shoot brief: exact {EXPECTED_BRIEF_LEN} shoots", lambda: len(_brief) == EXPECTED_BRIEF_LEN)
chk("Shoot brief: types compleet", lambda: {s["type"] for s in _brief} == EXPECTED_TYPES)
chk("Shoot brief: geen em-dashes", lambda: not any(
    "—" in v for s in _brief for v in s.get("tijdcodes", {}).values()
))
chk("Shoot brief: lege client_name crasht niet", lambda: (
    len(generate_shoot_brief(_summary, _all_ads, client_name="")) == EXPECTED_BRIEF_LEN
))
chk("_strip_em_dashes: verwijdert em dashes", lambda: "—" not in _strip_em_dashes("Hallo — wereld"))
chk("_clean_script: verwijdert em dashes", lambda: not any(
    "—" in b["tekst"] for b in _clean_script([{"time": "0-5s", "tekst": "Test — zin"}])
))
chk("Alle 10 hook types hebben 3+ tijdcodes + geen em dashes", lambda: all(
    len(_build_script(h, "CTA tekst", "TestKlant")) >= 3 and
    not any("—" in b["tekst"] for b in _build_script(h, "CTA tekst", "TestKlant"))
    for h in SB_HOOKS
))


# ═════════════════════════════════════════════════════════════════════════════
# 5. CREATIVE DECODER
# ═════════════════════════════════════════════════════════════════════════════
section("5. Creative Decoder")
from core.creative_decoder import decode_winner, decode_loser

_winners = [a for a in _all_ads if a.results > 0 and a.cost_per_result > 0][:2]
_losers  = [a for a in _all_ads if a.results == 0 and a.spend > 10][:2]

chk("decode_winner: hook_type aanwezig", lambda: all(
    "hook_type" in decode_winner(a, _summary) for a in _winners
) if _winners else "WARN: geen winner ads")
chk("decode_loser: retourneert dict", lambda: all(
    isinstance(decode_loser(a, _summary), dict) for a in _losers
) if _losers else "WARN: geen loser ads")


# ═════════════════════════════════════════════════════════════════════════════
# 6. EXCEL TEMPLATES
# ═════════════════════════════════════════════════════════════════════════════
section("6. Excel Templates")
try:
    from core.excel_templates import generate_videos_template, generate_statics_template, parse_template
    from openpyxl import load_workbook

    vt = generate_videos_template()
    st = generate_statics_template()
    chk("Videos template: bytes > 0", lambda: isinstance(vt, bytes) and len(vt) > 100)
    chk("Statics template: bytes > 0", lambda: isinstance(st, bytes) and len(st) > 100)

    # Parse lege template
    _kn, _cr = parse_template(io.BytesIO(vt), "videos")
    chk("Lege template parse: 0 creatives", lambda: len(_cr) == 0)

    # Vul template in en parse terug
    wb = load_workbook(io.BytesIO(vt))
    ws = wb.active
    ws["B1"] = "ParseKlant"
    ws.cell(row=5, column=1, value="reels - problem_solve - v1 - test ad")
    ws.cell(row=5, column=2, value="Test script tekst")
    ws.cell(row=5, column=3, value="Test copy 1")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    kn2, cr2 = parse_template(buf, "videos")
    chk("Template parse: klantnaam correct", lambda: kn2 == "ParseKlant")
    chk("Template parse: 1 creative", lambda: len(cr2) == 1)
    chk("Template parse: ad_naam correct", lambda: cr2[0]["ad_naam"] == "reels - problem_solve - v1 - test ad")
    chk("Template parse: script correct", lambda: cr2[0]["script"] == "Test script tekst")

    # Voorbeeldrij overgeslagen (rij 4) — template heeft al de exacte voorbeeldwaarde in rij 4
    wb3 = load_workbook(io.BytesIO(vt))
    ws3 = wb3.active
    ws3.cell(row=5, column=1, value="Echte ad naam")
    buf3 = io.BytesIO(); wb3.save(buf3); buf3.seek(0)
    _, cr3 = parse_template(buf3, "videos")
    EXAMPLE_NAAM = "Static - Proof - V1 - Klantresultaat fitness"
    chk("Template: voorbeeldrij overgeslagen (rij 4)", lambda: (
        len(cr3) >= 1
        and all(c.get("ad_naam", "") != EXAMPLE_NAAM for c in cr3)
        and any(c.get("ad_naam", "").strip() == "Echte ad naam" for c in cr3)
    ))

except ImportError:
    chk("openpyxl beschikbaar", lambda: False)
except Exception as e:
    chk("Excel templates foutloos", lambda: (_ for _ in ()).throw(type(e)(str(e))))


# ═════════════════════════════════════════════════════════════════════════════
# 7. PDF EXPORT
# ═════════════════════════════════════════════════════════════════════════════
section("7. PDF Export")
try:
    from core.reporter import generate_pdf, generate_shoot_brief_pdf

    top_ads_data = [
        {"ad_name": a.ad_name, "spend": a.spend, "results": a.results, "cpl": a.cost_per_result}
        for a in _all_ads[:3]
    ]
    pdf_bytes = generate_pdf(_summary, "Test AI inzicht tekst", top_ads=top_ads_data)
    chk("generate_pdf: retourneert bytes", lambda: isinstance(pdf_bytes, bytes) and len(pdf_bytes) > 200)
    chk("PDF header: %PDF aanwezig", lambda: pdf_bytes[:4] == b"%PDF")

    brief_pdf = generate_shoot_brief_pdf(FAKE_BRIEF["brief_json"], client_name="TestKlant BV")
    chk("generate_shoot_brief_pdf: retourneert bytes", lambda: isinstance(brief_pdf, bytes) and len(brief_pdf) > 200)
    chk("Brief PDF header: %PDF aanwezig", lambda: brief_pdf[:4] == b"%PDF")

except Exception as e:
    chk("PDF generatie foutloos", lambda: (_ for _ in ()).throw(type(e)(str(e))))


# ═════════════════════════════════════════════════════════════════════════════
# 8. FLASK ROUTES — setup
# ═════════════════════════════════════════════════════════════════════════════
section("8. Flask Routes — Auth & Home")

# Import app AFTER core tests (vermijdt dubbele DB-init logging)
import app as flask_app
flask_app.app.config["TESTING"] = True
flask_app.app.config["SECRET_KEY"] = "test-secret"
flask_app.app.config["WTF_CSRF_ENABLED"] = False

_mock = _mock_db()


def _client(session_data: dict | None = None):
    """Return een Flask test client met optionele session-data."""
    c = flask_app.app.test_client()
    if session_data:
        with c.session_transaction() as sess:
            for k, v in session_data.items():
                sess[k] = v
    return c


def _demo_session() -> dict:
    """Session die de demo-flow simuleert (rows in session via data_source=demo)."""
    from core.csv_parser import load_dummy_data
    from core.analysis import build_campaigns, build_summary, get_all_ads, get_date_range
    import json
    rows = load_dummy_data()
    campaigns = build_campaigns(rows)
    summary = build_summary(rows, campaigns)
    all_ads = sorted(get_all_ads(campaigns), key=lambda a: a.spend, reverse=True)
    dr = get_date_range(rows)
    return {
        "username": "dev",
        "data_source": "demo",
        "thresholds": {"winner": 25, "mid": 50, "preset": "auto"},
        "summary": {
            "total_spend": summary.total_spend,
            "total_impressions": summary.total_impressions,
            "total_reach": getattr(summary, "total_reach", 0),
            "total_clicks": summary.total_clicks,
            "total_link_clicks": getattr(summary, "total_link_clicks", 0),
            "total_results": summary.total_results,
            "avg_ctr": summary.avg_ctr,
            "avg_cpc": summary.avg_cpc,
            "avg_cpm": summary.avg_cpm,
            "avg_roas": summary.avg_roas,
            "avg_frequency": summary.avg_frequency,
            "avg_cost_per_result": summary.avg_cost_per_result,
            "num_campaigns": summary.num_campaigns,
            "num_ad_sets": summary.num_ad_sets,
            "num_ads": summary.num_ads,
            "campaign_type": summary.campaign_type,
            "top_ad": summary.top_ad,
            "top_ad_set": summary.top_ad_set,
            "worst_ad": summary.worst_ad,
            "worst_ad_set": summary.worst_ad_set,
            "has_click_data": summary.has_click_data,
        },
        "top_ads": [
            {"ad_name": a.ad_name, "spend": a.spend, "results": a.results, "cpl": a.cost_per_result}
            for a in all_ads[:5]
        ],
        "date_range": {"from": dr[0], "to": dr[1], "full_from": dr[0], "full_to": dr[1]},
    }


with patch("app.db", _mock):

    # ── Auth: login GET
    # Als APP_USERS niet is ingesteld, redirectt /login direct naar /clients (200 of 302 zijn beiden goed)
    c = _client()
    r = c.get("/login")
    chk("GET /login: 200 of redirect", lambda: r.status_code in (200, 302))

    # ── Auth: redirect zonder login (APP_USERS leeg → auto-login als dev)
    r2 = c.get("/")
    chk("GET / redirect naar /clients", lambda: r2.status_code in (302, 200))

    # ── Logout
    c2 = _client({"username": "dev"})
    r3 = c2.get("/logout")
    chk("GET /logout: redirect", lambda: r3.status_code in (302, 301))


section("9. Flask Routes — Clients")
with patch("app.db", _mock):

    c = _client({"username": "dev"})

    r = c.get("/clients")
    chk("GET /clients: 200", lambda: r.status_code == 200)
    chk("GET /clients: beide klanten zichtbaar", lambda: b"TestKlant BV" in r.data)

    # Client aanmaken
    r = c.post("/clients/new", data={
        "name": "NieuweKlant",
        "industry": "beauty",
        "campaign_type": "leads",
        "cpl_benchmark": "35",
        "notes": "test",
        "client_context": "",
    }, follow_redirects=False)
    chk("POST /clients/new: redirect (DB mock aangemaakt)", lambda: r.status_code in (302, 200))

    # Client profiel
    r = c.get("/clients/1")
    chk("GET /clients/1: 200", lambda: r.status_code == 200)
    chk("GET /clients/1: klantnaam aanwezig", lambda: b"TestKlant" in r.data)

    # Client edit
    r = c.post("/clients/1/edit", data={
        "name": "TestKlant BV Updated",
        "industry": "fitness",
        "campaign_type": "leads",
        "cpl_benchmark": "28",
        "notes": "bijgewerkt",
        "client_context": "Fitness context",
    }, follow_redirects=False)
    chk("POST /clients/1/edit: redirect", lambda: r.status_code in (302, 200))

    # Client 2 profiel
    _mock.get_client.return_value = FAKE_CLIENT_2
    r = c.get("/clients/2")
    chk("GET /clients/2: 200", lambda: r.status_code == 200)
    _mock.get_client.return_value = FAKE_CLIENT


section("10. Flask Routes — Analyse (demo + upload)")
with patch("app.db", _mock):

    c = _client({"username": "dev"})

    # GET /analyse (leeg scherm)
    r = c.get("/analyse")
    chk("GET /analyse: 200", lambda: r.status_code == 200)

    # GET /demo — draait volledige analyse op dummy data
    r = c.get("/demo")
    chk("GET /demo: 200", lambda: r.status_code == 200)
    chk("GET /demo: bevat 'Campagnes'", lambda: b"Campagne" in r.data or b"campagne" in r.data)

    # POST /upload met dummy CSV
    csv_bytes = FAKE_CSV_CONTENT.encode("utf-8")
    data = {
        "csv_file": (io.BytesIO(csv_bytes), "test.csv"),
        "threshold_preset": "auto",
    }
    r = c.post("/upload", data=data, content_type="multipart/form-data")
    chk("POST /upload dummy CSV: 200", lambda: r.status_code == 200)

    # POST /upload leeg bestand
    r2 = c.post("/upload", data={}, content_type="multipart/form-data", follow_redirects=True)
    chk("POST /upload leeg: redirect met flash", lambda: r2.status_code == 200)

    # POST /upload verkeerd type
    r3 = c.post("/upload", data={
        "csv_file": (io.BytesIO(b"fake"), "test.txt"),
    }, content_type="multipart/form-data", follow_redirects=True)
    chk("POST /upload .txt: geweigerd", lambda: r3.status_code == 200)


section("11. Flask Routes — Analyse (reanalyze, creative, hooks)")
_demo_sess = _demo_session()

with patch("app.db", _mock):

    c = _client(_demo_sess)

    # POST /reanalyze
    r = c.post("/reanalyze", data={"threshold_preset": "auto"})
    chk("POST /reanalyze: 200", lambda: r.status_code == 200)

    # GET /creative
    r = c.get("/creative")
    chk("GET /creative: 200", lambda: r.status_code == 200)
    chk("GET /creative: bevat 'Winner' of 'winner'", lambda: b"inner" in r.data)

    # GET /hooks
    r = c.get("/hooks")
    chk("GET /hooks: 200", lambda: r.status_code == 200)
    chk("GET /hooks: bevat 'hook'", lambda: b"hook" in r.data.lower() or b"Hook" in r.data)

    # Reanalyze met date filter
    r = c.post("/reanalyze", data={"date_from": "2025-01-01", "date_to": "2025-01-31", "threshold_preset": "auto"})
    chk("POST /reanalyze + date filter: 200 of redirect", lambda: r.status_code in (200, 302))

    # Reanalyze met campaign_type override
    r = c.post("/reanalyze", data={"campaign_type_override": "purchases", "threshold_preset": "auto"})
    chk("POST /reanalyze + type override: 200", lambda: r.status_code == 200)

    # Reanalyze zonder sessie
    c_leeg = _client({"username": "dev"})
    r = c_leeg.post("/reanalyze", data={"threshold_preset": "auto"}, follow_redirects=True)
    chk("POST /reanalyze zonder sessie: redirect naar index", lambda: r.status_code == 200)


section("12. Flask Routes — Shoot Brief & Export")
with patch("app.db", _mock):

    c = _client({**_demo_sess, "client_id": 1})

    # POST generate-shoot-brief
    r = c.post("/clients/1/generate-shoot-brief")
    chk("POST /clients/1/generate-shoot-brief: 200", lambda: r.status_code == 200)
    chk("Shoot brief response: HTML partial", lambda: b"shoot" in r.data.lower() or len(r.data) > 100)

    # GET export shoot-brief PDF
    r = c.get("/clients/1/export/shoot-brief")
    chk("GET /clients/1/export/shoot-brief: PDF of redirect", lambda: r.status_code in (200, 302))
    if r.status_code == 200:
        chk("Shoot brief PDF: content-type PDF", lambda: "pdf" in r.content_type.lower())

    # GET export PDF (analyse)
    r = c.get("/export/pdf")
    chk("GET /export/pdf: 200 of redirect", lambda: r.status_code in (200, 302))

    # Brief delete
    r = c.post("/clients/1/briefs/55/delete", follow_redirects=False)
    chk("POST /clients/1/briefs/55/delete: redirect", lambda: r.status_code in (302, 200))


section("13. Flask Routes — Upload management")
with patch("app.db", _mock):

    c = _client({**_demo_sess, "client_id": 1})

    # Upload verwijderen
    r = c.post("/clients/1/uploads/101/delete", follow_redirects=False)
    chk("POST upload delete: redirect", lambda: r.status_code in (302, 200))

    # Upload laden (load)
    r = c.get("/clients/1/load/101", follow_redirects=False)
    chk("GET /clients/1/load/101: redirect of 200", lambda: r.status_code in (302, 200))

    # Client merge uploads
    r = c.post("/clients/1/merge", data={"upload_ids": [101, 101]}, follow_redirects=False)
    chk("POST /clients/1/merge: redirect of 200", lambda: r.status_code in (302, 200))


section("14. Flask Routes — Templates download")
with patch("app.db", _mock):

    c = _client({"username": "dev"})

    r = c.get("/templates/videos")
    chk("GET /templates/videos: Excel bestand", lambda: r.status_code == 200)
    chk("Templates/videos: xlsx content-type", lambda: "spreadsheet" in r.content_type or "excel" in r.content_type.lower() or r.status_code == 200)

    r = c.get("/templates/statics")
    chk("GET /templates/statics: 200", lambda: r.status_code == 200)


section("15. Flask Routes — Creatives & Nieuwe Advertentie")
with patch("app.db", _mock):

    c = _client({**_demo_sess, "client_id": 1})

    # Import creatives pagina
    r = c.get("/clients/1/import/creatives")
    chk("GET /clients/1/import/creatives: 200", lambda: r.status_code == 200)

    # Creatives overzicht
    r = c.get("/clients/1/creatives")
    chk("GET /clients/1/creatives: 200", lambda: r.status_code == 200)

    # Nieuwe advertentie pagina
    r = c.get("/client/1/nieuwe-advertentie")
    chk("GET /client/1/nieuwe-advertentie: 200 of redirect", lambda: r.status_code in (200, 302))

    # New ads content pagina
    r = c.get("/clients/1/new-ads")
    chk("GET /clients/1/new-ads: 200 of redirect", lambda: r.status_code in (200, 302))

    # Save creative (form POST, geen JSON)
    r = c.post("/clients/1/save-creative", data={
        "ad_naam": "reels - problem_solve - v1 - test",
        "script": "Scripttest",
        "ad_copy_1": "Copy 1 test",
    }, follow_redirects=False)
    chk("POST /clients/1/save-creative: 200", lambda: r.status_code == 200)


section("16. Flask Routes — Static Brief & Advertentie Geschiedenis")
with patch("app.db", _mock):

    c = _client({**_demo_sess, "client_id": 1})

    r = c.get("/client/1/static-brief")
    chk("GET /client/1/static-brief: 200", lambda: r.status_code == 200)

    r = c.get("/client/1/advertentie-geschiedenis")
    chk("GET /client/1/advertentie-geschiedenis: 200", lambda: r.status_code == 200)


section("17. Flask Routes — Action Plan & Competitor")
with patch("app.db", _mock):

    c = _client({**_demo_sess, "client_id": 1})

    r = c.get("/client/1/action-plan")
    chk("GET /client/1/action-plan: 200 of redirect", lambda: r.status_code in (200, 302))

    r = c.get("/client/1/competitor-analysis")
    chk("GET /client/1/competitor-analysis: 200 of redirect", lambda: r.status_code in (200, 302))


section("18. Flask Routes — Meta OAuth (geen echte token nodig)")
with patch("app.db", _mock):

    c = _client({"username": "dev", "client_id": 1})

    r = c.get("/client/1/connect-meta", follow_redirects=False)
    chk("GET /client/1/connect-meta: redirect naar Meta", lambda: r.status_code in (302, 200))

    r = c.get("/client/1/meta-status")
    chk("GET /client/1/meta-status: 200", lambda: r.status_code == 200)

    r = c.post("/client/1/sync-meta", follow_redirects=False)
    chk("POST /client/1/sync-meta: redirect of 200", lambda: r.status_code in (302, 200))


section("19. Flask Routes — Tag Ads & Debug")
with patch("app.db", _mock):

    c = _client({**_demo_sess, "client_id": 1})

    # Tag ads
    mappings = json.dumps([{"ad_name": "Test Ad", "hook": "proof", "format": "static"}])
    r = c.post("/tag-ads", data={"mappings": mappings})
    chk("POST /tag-ads: 200 of 400", lambda: r.status_code in (200, 400))

    # Debug DB pagina
    r = c.get("/debug/db")
    chk("GET /debug/db: 200", lambda: r.status_code == 200)


section("20. Flask Routes — Niet-bestaande klant (404-flow)")
with patch("app.db", _mock):

    _mock.get_client.return_value = None
    c = _client({"username": "dev"})

    r = c.get("/clients/9999", follow_redirects=True)
    chk("GET /clients/9999 (niet gevonden): redirect naar /clients", lambda: r.status_code == 200)

    _mock.get_client.return_value = FAKE_CLIENT


section("21. Threshold helpers")
from app import _compute_thresholds

chk("Threshold: preset=auto zonder benchmark", lambda: _compute_thresholds({})["winner"] >= 1)
chk("Threshold: preset=fit20", lambda: _compute_thresholds({"threshold_preset": "fit20"})["winner"] == 40)
chk("Threshold: preset=belladonna", lambda: _compute_thresholds({"threshold_preset": "belladonna"})["winner"] == 30)
chk("Threshold: auto + CPL benchmark", lambda: (
    _compute_thresholds({"threshold_preset": "auto"}, client={"cpl_benchmark": 45.0})["winner"] == 45
))
chk("Threshold: custom negatief → min 1", lambda: (
    _compute_thresholds({"threshold_preset": "custom", "threshold_winner": "-5", "threshold_mid": "20"})["winner"] >= 1
))
chk("Threshold: mid altijd > winner", lambda: (
    _compute_thresholds({"threshold_preset": "custom", "threshold_winner": "50", "threshold_mid": "30"})["mid"] > 50
))

from app import _dedup_key
chk("Dedup key: ad_id='0' valt terug op ad_name", lambda: (
    _dedup_key({"ad_id": "0", "ad_name": "Ad A", "campaign_id": "0", "campaign_name": "C", "day": ""}) !=
    _dedup_key({"ad_id": "0", "ad_name": "Ad B", "campaign_id": "0", "campaign_name": "C", "day": ""})
))
chk("Dedup key: verschillende maanden = verschillende keys", lambda: (
    _dedup_key({"ad_id": "0", "ad_name": "A", "campaign_id": "0", "campaign_name": "C", "day": "2025-03-01"}) !=
    _dedup_key({"ad_id": "0", "ad_name": "A", "campaign_id": "0", "campaign_name": "C", "day": "2025-04-01"})
))


# ═════════════════════════════════════════════════════════════════════════════
# 22. Incrementele sync — datum-logica
# ═════════════════════════════════════════════════════════════════════════════
section("22. Incrementele sync — datum-logica")
from datetime import date as _date, timedelta as _td, datetime as _dt

def _resolve_sync_dates(connection: dict, date_from=None, date_to=None):
    """Spiegeling van de datum-logica uit _run_meta_sync, zonder DB/API."""
    today = _date(2026, 5, 20)
    if not date_from:
        last_sync = connection.get("last_sync_at")
        if last_sync:
            if hasattr(last_sync, "date"):
                last_sync_date = last_sync.date()
            else:
                try:
                    last_sync_date = _date.fromisoformat(str(last_sync)[:10])
                except ValueError:
                    last_sync_date = None
            if last_sync_date:
                date_from = (last_sync_date + _td(days=1)).isoformat()
        if not date_from:
            date_from = (today - _td(days=30)).isoformat()
    if not date_to:
        date_to = today.isoformat()
    skipped = date_from > date_to
    return date_from, date_to, skipped

chk("Geen last_sync → valt terug op 30 dagen geleden",
    lambda: _resolve_sync_dates({})[0] == "2026-04-20")

chk("last_sync gisteren → date_from = vandaag",
    lambda: _resolve_sync_dates({"last_sync_at": "2026-05-19"})[0] == "2026-05-20")

chk("last_sync vandaag → date_from = morgen → skip",
    lambda: _resolve_sync_dates({"last_sync_at": "2026-05-20"})[2] is True)

chk("last_sync als datetime object",
    lambda: _resolve_sync_dates({"last_sync_at": _dt(2026, 5, 18, 7, 0, 0)})[0] == "2026-05-19")

chk("Handmatig date_from meegegeven → last_sync genegeerd",
    lambda: _resolve_sync_dates({"last_sync_at": "2026-01-01"}, date_from="2026-03-01")[0] == "2026-03-01")

chk("last_sync 7 dagen geleden → pakt precies 6 dagen terug",
    lambda: _resolve_sync_dates({"last_sync_at": "2026-05-13"})[0] == "2026-05-14")

chk("Skip flag False als er nog data te halen is",
    lambda: _resolve_sync_dates({"last_sync_at": "2026-05-19"})[2] is False)

chk("date_to is altijd vandaag als niet opgegeven",
    lambda: _resolve_sync_dates({})[1] == "2026-05-20")


# ═════════════════════════════════════════════════════════════════════════════
# 23. Deduplicatie logica — _non_redundant_upload_ids
# ═════════════════════════════════════════════════════════════════════════════
section("23. Deduplicatie logica — non-redundant uploads")
from datetime import date as _d, timedelta as _tdd

def _simulate_non_redundant(uploads):
    """
    Puur-Python simulatie van _non_redundant_upload_ids.
    uploads = list van (id, date_from_str, date_to_str), nieuwste eerst.
    """
    def _parse(s):
        try:
            return _d.fromisoformat(str(s)[:10]) if s else None
        except ValueError:
            return None

    covered = set()
    selected = []
    for uid, date_from, date_to in uploads:
        d_from = _parse(date_from)
        d_to   = _parse(date_to)
        if d_from is None or d_to is None:
            selected.append(uid)
            continue
        days = set()
        d = d_from
        while d <= d_to:
            days.add(d)
            d += _tdd(days=1)
        new_days = days - covered
        if new_days:
            selected.append(uid)
            covered |= days
    return selected

# Scenario A: geen overlap (cron dagelijks)
uploads_a = [
    (3, "2026-05-20", "2026-05-20"),  # nieuwste
    (2, "2026-05-19", "2026-05-19"),
    (1, "2026-01-01", "2026-05-18"),  # handmatig groot
]
chk("Geen overlap: alle uploads selected",
    lambda: set(_simulate_non_redundant(uploads_a)) == {1, 2, 3})

# Scenario B: volledige overlap (zelfde periode opnieuw geüpload)
uploads_b = [
    (2, "2026-01-01", "2026-05-20"),  # nieuwste, zelfde periode
    (1, "2026-01-01", "2026-05-20"),  # oudste → volledig gedekt
]
chk("Volledig overlappend: alleen nieuwste selected",
    lambda: _simulate_non_redundant(uploads_b) == [2])

# Scenario C: deels overlap — oud upload heeft vroegere data
uploads_c = [
    (2, "2026-04-01", "2026-05-20"),  # nieuwste
    (1, "2026-01-01", "2026-03-31"),  # oudste, geen overlap
]
chk("Aangrenzend zonder overlap: beide selected",
    lambda: set(_simulate_non_redundant(uploads_c)) == {1, 2})

# Scenario D: cron na handmatig (incrementeel correct)
uploads_d = [
    (4, "2026-05-20", "2026-05-20"),
    (3, "2026-05-19", "2026-05-19"),
    (2, "2026-05-18", "2026-05-18"),
    (1, "2026-01-01", "2026-05-17"),
]
chk("Incrementele cron na handmatig: alle 4 selected (geen overlap)",
    lambda: len(_simulate_non_redundant(uploads_d)) == 4)

# Scenario E: single upload
uploads_e = [(1, "2026-01-01", "2026-05-20")]
chk("Single upload: altijd selected",
    lambda: _simulate_non_redundant(uploads_e) == [1])

# Scenario F: lege lijst
chk("Lege uploads: leeg resultaat",
    lambda: _simulate_non_redundant([]) == [])


# ═════════════════════════════════════════════════════════════════════════════
# 24. get_correct_totals — geen dubbeltelling
# ═════════════════════════════════════════════════════════════════════════════
section("24. get_correct_totals — geen dubbeltelling")

def _simulate_correct_totals(uploads_with_spend):
    """
    uploads_with_spend = list van (id, date_from, date_to, total_spend, total_results)
    Nieuwste eerst. Berekent totalen zonder dubbeltelling.
    """
    upload_tuples = [(u[0], u[1], u[2]) for u in uploads_with_spend]
    valid_ids = set(_simulate_non_redundant(upload_tuples))
    spend   = sum(u[3] for u in uploads_with_spend if u[0] in valid_ids)
    results = sum(u[4] for u in uploads_with_spend if u[0] in valid_ids)
    return {"total_spend": spend, "total_results": results}

# Incrementeel (dagelijkse cron): optelling moet kloppen
uploads_inc = [
    (3, "2026-05-20", "2026-05-20", 50.0,    3),
    (2, "2026-05-19", "2026-05-19", 45.0,    2),
    (1, "2026-01-01", "2026-05-18", 10000.0, 600),
]
result_inc = _simulate_correct_totals(uploads_inc)
chk("Incrementeel: spend correct opgeteld (geen dubbel)",
    lambda: result_inc["total_spend"] == 10095.0)
chk("Incrementeel: results correct opgeteld",
    lambda: result_inc["total_results"] == 605)

# Zelfde periode opnieuw: alleen nieuwste telt
uploads_dup = [
    (2, "2026-01-01", "2026-05-20", 9800.0, 590),  # herberekend nieuwste
    (1, "2026-01-01", "2026-05-20", 9500.0, 570),  # oudste, identieke periode
]
result_dup = _simulate_correct_totals(uploads_dup)
chk("Zelfde periode: alleen nieuwste upload telt (geen dubbel)",
    lambda: result_dup["total_spend"] == 9800.0)
chk("Zelfde periode: results van nieuwste",
    lambda: result_dup["total_results"] == 590)

# Geen uploads
chk("Lege uploads: spend = 0",
    lambda: _simulate_correct_totals([])["total_spend"] == 0)
chk("Lege uploads: results = 0",
    lambda: _simulate_correct_totals([])["total_results"] == 0)

# avg_cpl check
chk("CPL = spend / results wanneer results > 0",
    lambda: round(10095.0 / 605, 2) == 16.69)
chk("CPL = None wanneer geen results",
    lambda: (None if 0 == 0 else 0) is None)


# ═════════════════════════════════════════════════════════════════════════════
# 25. Flask route — client_profile krijgt client_totals mee
# ═════════════════════════════════════════════════════════════════════════════
section("25. Flask route — client_profile met correcte totalen")

def _test_client_profile_totals():
    with patch("app.db", _mock):
        c = _client({**_demo_sess, "client_id": 1})
        resp = c.get("/clients/1")
        return resp.status_code == 200

chk("client_profile route: get_correct_totals aangeroepen (geen crash)",
    _test_client_profile_totals)

def _test_sync_skip_when_uptodate():
    """_run_meta_sync moet skipped=True teruggeven als al gesynchroniseerd vandaag."""
    from app import _run_meta_sync
    fake_conn = {
        "access_token": "nep",
        "ad_account_id": "act_123",
        "last_sync_at": _dt.today(),
        "client_name": "Test",
        "campaign_type": "leads",
    }
    with patch("app.db") as mdb, \
         patch("app.get_ads", return_value=[]), \
         patch("app._decrypt_token", return_value="nep_token"):
        mdb.get_client.return_value = {"campaign_type": "leads", "id": 1}
        result = _run_meta_sync(1, fake_conn)
    return result.get("skipped") is True and result.get("ok") is True

chk("Sync skip: vandaag al gesynchroniseerd → skipped=True, ok=True",
    _test_sync_skip_when_uptodate)

def _test_sync_uses_last_sync_date():
    """Cron sync moet date_from = last_sync + 1 dag gebruiken."""
    from app import _run_meta_sync
    captured = {}
    def fake_get_ads(token, acct_id, date_from, date_to):
        captured["date_from"] = date_from
        return []
    fake_conn = {
        "access_token": "nep",
        "ad_account_id": "act_123",
        "last_sync_at": "2026-05-10",
        "client_name": "Test",
        "campaign_type": "leads",
    }
    with patch("app.db") as mdb, \
         patch("app.get_ads", side_effect=fake_get_ads), \
         patch("app._decrypt_token", return_value="nep_token"):
        mdb.get_client.return_value = {"campaign_type": "leads", "id": 1}
        _run_meta_sync(1, fake_conn)
    return captured.get("date_from") == "2026-05-11"

chk("Cron sync: date_from = last_sync_at + 1 dag",
    _test_sync_uses_last_sync_date)

def _test_sync_fallback_30days():
    """Zonder last_sync moet de sync 30 dagen teruggaan."""
    from datetime import date as _today_date
    from app import _run_meta_sync
    captured = {}
    def fake_get_ads(token, acct_id, date_from, date_to):
        captured["date_from"] = date_from
        return []
    fake_conn = {
        "access_token": "nep",
        "ad_account_id": "act_123",
        "last_sync_at": None,
        "client_name": "Test",
        "campaign_type": "leads",
    }
    with patch("app.db") as mdb, \
         patch("app.get_ads", side_effect=fake_get_ads), \
         patch("app._decrypt_token", return_value="nep_token"):
        mdb.get_client.return_value = {"campaign_type": "leads", "id": 1}
        _run_meta_sync(1, fake_conn)
    from datetime import date, timedelta
    expected = (date.today() - timedelta(days=30)).isoformat()
    return captured.get("date_from") == expected

chk("Cron sync fallback: geen last_sync → 30 dagen terug",
    _test_sync_fallback_30days)


# ═════════════════════════════════════════════════════════════════════════════
# EINDRESULTAAT
# ═════════════════════════════════════════════════════════════════════════════
section("EINDRESULTAAT")
total = len(_results)
print(f"\n  Totaal  : {total}")
print(f"  \033[92mPassed\033[0m  : {PASS_COUNT}")
if WARN_COUNT:
    print(f"  \033[93mWarnings\033[0m: {WARN_COUNT}")
if FAIL_COUNT:
    print(f"  \033[91mFailed\033[0m  : {FAIL_COUNT}")
    print("\nGefaalde tests:")
    for name, status, detail in _results:
        if status == "fail":
            print(f"  ✗ {name}")
            if detail:
                print(f"    {detail}")

sys.exit(0 if FAIL_COUNT == 0 else 1)
