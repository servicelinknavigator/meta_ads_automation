import os
import sys
import re
import csv as _csv_module
import json
import logging
import hashlib
import hmac
from pathlib import Path
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from markupsafe import Markup, escape as html_escape
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from datetime import timedelta
import io

load_dotenv()
sys.path.insert(0, str(Path(__file__).parent))

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

from core.csv_parser import parse_csv, parse_csv_string, load_dummy_data, validate_csv
from core.analysis import (
    build_campaigns, build_summary, build_ad_chart_data, get_all_ads,
    get_date_range, filter_rows_by_date, build_wow_comparison,
    filter_zero_spend, build_ad_delivery_map,
)
from core.generation import generate_insights
from core.reporter import generate_pdf, generate_shoot_brief_pdf
from core.creative_decoder import decode_winner, decode_loser
from core.axes_mapper import map_axes
from core.smart_generator import generate_testkit
from core.hook_analyzer import (
    aggregate_hook_performance, aggregate_format_performance,
    get_winning_combinations, get_untested_hooks, get_untested_formats,
    get_unknown_ads,
)
from core.ai_client import suggest_ad_tags
from core.shoot_brief import generate_shoot_brief
from core.excel_templates import generate_videos_template, generate_statics_template, parse_template
import core.db as db

# ── In-memory creative cache ───────────────────────────────────────────────────
import threading as _threading
_CREATIVE_CACHE: dict = {}
_CREATIVE_CACHE_MAX = 120
_CREATIVE_CACHE_LOCK = _threading.Lock()


def _cache_key(ad) -> str:
    s = f"{ad.ad_name}:{ad.campaign_name}:{ad.spend:.2f}:{ad.results}"
    return hashlib.sha256(s.encode()).hexdigest()[:16]


def _cache_set(key: str, value: dict) -> None:
    with _CREATIVE_CACHE_LOCK:
        if len(_CREATIVE_CACHE) >= _CREATIVE_CACHE_MAX:
            _CREATIVE_CACHE.pop(next(iter(_CREATIVE_CACHE)))
        _CREATIVE_CACHE[key] = value


def _cache_get(key: str) -> dict | None:
    with _CREATIVE_CACHE_LOCK:
        return _CREATIVE_CACHE.get(key)


# ── App setup ──────────────────────────────────────────────────────────────────
app = Flask(__name__)
_flask_secret = os.getenv("FLASK_SECRET_KEY", "")
if not _flask_secret:
    logger.warning("FLASK_SECRET_KEY not set — using insecure default. Set this in production!")
    _flask_secret = "dev-secret-change-me"
app.secret_key = _flask_secret
app.config["MAX_CONTENT_LENGTH"] = int(os.getenv("MAX_UPLOAD_MB", 50)) * 1024 * 1024
app.permanent_session_lifetime = timedelta(hours=4)

if not os.getenv("TOKEN_ENCRYPTION_KEY"):
    logger.warning("TOKEN_ENCRYPTION_KEY not set — Meta OAuth tokens will be stored in plaintext!")

UPLOAD_FOLDER = Path(__file__).parent / "uploads"
UPLOAD_FOLDER.mkdir(exist_ok=True)

# Init DB schema on startup
with app.app_context():
    try:
        db.init_schema()
    except Exception as e:
        logger.warning("DB init skipped: %s", e)


# ── Auth ───────────────────────────────────────────────────────────────────────

def _get_users() -> dict:
    """Parse APP_USERS=dominique:pass1,rick:pass2 from env."""
    raw = os.getenv("APP_USERS", "")
    users = {}
    for part in raw.split(","):
        part = part.strip()
        if ":" in part:
            u, p = part.split(":", 1)
            users[u.strip().lower()] = p.strip()
    return users


def _auth_enabled() -> bool:
    return bool(os.getenv("APP_USERS", "").strip())


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if _auth_enabled() and not session.get("username"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# ── Helpers ────────────────────────────────────────────────────────────────────

def _dedup_key(row: dict) -> tuple:
    """
    Build a dedup key for merge operations.
    Falls back to name when the ID column is absent or defaulted to '0'.
    This handles both day-by-day exports (have real ad_id + day) and
    summary-format exports (no ad_id/campaign_id/day column → all default to '0'/empty).
    """
    ad_key = str(row.get("ad_id", "") or "")
    if not ad_key or ad_key == "0":
        ad_key = row.get("ad_name", "")

    camp_key = str(row.get("campaign_id", "") or "")
    if not camp_key or camp_key == "0":
        camp_key = row.get("campaign_name", "")

    return (ad_key, camp_key, row.get("day", ""))


def _load_name_overrides() -> dict:
    """Load client's saved ad name mappings from DB for current session client."""
    client_id = session.get("client_id")
    if client_id and db.is_available():
        try:
            return db.get_ad_name_mappings(client_id)
        except Exception:
            pass
    return {}


def _compute_thresholds(form=None, client=None) -> dict:
    preset = (form.get("threshold_preset", "auto") if form else "auto")
    if preset == "fit20":
        return {"winner": 40, "mid": 60, "preset": "fit20"}
    if preset == "belladonna":
        return {"winner": 30, "mid": 50, "preset": "belladonna"}
    if preset == "custom" and form:
        try:
            w = max(1, int(form.get("threshold_winner", 30)))
            m = max(w + 1, int(form.get("threshold_mid", 50)))
            return {"winner": w, "mid": m, "preset": "custom"}
        except (ValueError, TypeError):
            pass
    # Auto: use client CPL benchmark if available
    if preset == "auto" and client and client.get("cpl_benchmark"):
        w = max(1, int(client["cpl_benchmark"]))
        m = max(w + 1, int(client["cpl_benchmark"] * 1.5))
        return {"winner": w, "mid": m, "preset": "auto", "from_benchmark": True}
    return {"winner": 25, "mid": 50, "preset": "auto"}


_LOW_SPEND_THRESHOLD    = 25.0   # ads onder dit bedrag → uitgesloten van analyse
_EARLY_SIGNAL_THRESHOLD = 50.0   # €25-50 → "Vroeg signaal", geen winner/loser


def _process_df(rows: list, campaign_type_override: str = "",
                date_from: str = "", date_to: str = "",
                csv_content: str | None = None,
                name_overrides: dict | None = None,
                skip_db_save: bool = False) -> dict:
    valid, errors = validate_csv(rows)
    if not valid:
        return {"error": "; ".join(errors)}

    # Bouw delivery-map VOOR spend=0 filter (inactive ads kunnen spend=0 hebben)
    ad_delivery_map = build_ad_delivery_map(rows)

    # Filter ads zonder spend volledig uit — worden nergens getoond of getagd
    rows = filter_zero_spend(rows)
    if not rows:
        return {"error": "Alle advertenties hebben €0 spend — upload een CSV met actieve advertentie-data."}

    full_from, full_to = get_date_range(rows)
    if date_from or date_to:
        rows = filter_rows_by_date(rows, date_from, date_to)
        if not rows:
            return {"error": "Geen data gevonden voor de geselecteerde periode."}

    date_range = None
    if full_from:
        cur_from, cur_to = get_date_range(rows)
        date_range = {
            "from": cur_from or full_from,
            "to":   cur_to   or full_to,
            "full_from": full_from,
            "full_to":   full_to,
        }

    wow = build_wow_comparison(rows)
    campaigns = build_campaigns(rows)
    summary = build_summary(rows, campaigns, campaign_type_override=campaign_type_override)
    ad_chart_data = build_ad_chart_data(campaigns, summary.campaign_type)
    all_ads = sorted(get_all_ads(campaigns), key=lambda a: a.spend, reverse=True)

    # Haal creative content en cross-client patronen op voor AI context
    _client_id_for_ai = session.get("client_id")
    _ad_creatives_for_ai = {}
    _cross_client_for_ai = []
    if _client_id_for_ai and db.is_available():
        try:
            _ad_creatives_for_ai = db.get_ad_creatives(_client_id_for_ai)
        except Exception:
            pass
        try:
            _client_data = db.get_client(_client_id_for_ai)
            _industry = (_client_data or {}).get("industry", "")
            if _industry:
                _cross_client_for_ai = db.get_industry_cross_client_data(
                    _industry, exclude_client_id=_client_id_for_ai
                )
        except Exception:
            pass

    insights = generate_insights(summary, all_ads,
                                 ad_creatives=_ad_creatives_for_ai,
                                 cross_client_data=_cross_client_for_ai)

    _INACTIVE_STATUSES = {
        # Engels (Meta EN export)
        "inactive", "not_delivering", "not delivering", "paused", "disabled",
        "off", "stopped", "completed", "deleted", "archived", "rejected",
        # Nederlands (Meta NL export)
        "uitgeschakeld", "gepauzeerd", "niet actief", "niet_actief",
        "gestopt", "verwijderd", "afgewezen", "gearchiveerd",
        "niet geleverd", "niet_geleverd",
    }

    # Ads met te weinig spend krijgen een label maar geen urgentie-melding
    low_data_ads      = {a.ad_name for a in all_ads if 0 < a.spend < _LOW_SPEND_THRESHOLD}
    early_signal_ads  = {a.ad_name for a in all_ads if _LOW_SPEND_THRESHOLD <= a.spend < _EARLY_SIGNAL_THRESHOLD}

    _ACTIVE_STATUSES = {"active", "actief", "learning", "lerende fase", "in leer", "learning limited"}

    urgent_actions = []
    for a in all_ads:
        # Gebruik delivery_status van de Ad zelf, dan de delivery-map, dan leeg (onbekend).
        # Nooit "active" als default — als er geen statuskolom in de CSV zit weten we het niet.
        delivery = (a.delivery_status or "").lower().strip() or ad_delivery_map.get(a.ad_name, "").lower()
        is_inactive = delivery in _INACTIVE_STATUSES
        is_confirmed_active = delivery in _ACTIVE_STATUSES
        is_low_data = a.ad_name in low_data_ads

        if is_inactive or is_low_data:
            continue  # geen urgentiemeldingen voor inactieve of data-arme ads

        # Burning: altijd melden als ad actief of onbekend is — geld brandt al
        if a.results == 0 and a.spend > 50 and not is_inactive:
            urgent_actions.append({
                "type": "burning", "ad_name": a.ad_name,
                "ad_set_name": a.ad_set_name, "spend": round(a.spend),
            })
        # Fatigue: alleen melden als status BEVESTIGD actief is — heeft geen zin voor gestopte ads
        elif a.frequency > 3.5 and a.results > 0 and is_confirmed_active:
            urgent_actions.append({
                "type": "fatigue", "ad_name": a.ad_name,
                "ad_set_name": a.ad_set_name, "frequency": round(a.frequency, 1),
            })

    all_ads_json = json.dumps([{
        "ad_name":       a.ad_name,
        "ad_set_name":   a.ad_set_name,
        "campaign_name": a.campaign_name,
        "spend":         a.spend,
        "cpl":           a.cost_per_result,
        "ctr":           a.ctr,
        "results":       a.results,
        "cpc":           a.cpc,
        "roas":          a.roas,
        "impressions":   a.impressions,
        "cpm":           a.cpm,
        "frequency":        a.frequency,
        "delivery_status":  a.delivery_status,
    } for a in all_ads])

    # Limit top_ads to 5 entries to avoid session cookie overflow (4 KB browser limit)
    session["top_ads"] = [{
        "ad_name":         a.ad_name,
        "ad_set_name":     a.ad_set_name,
        "campaign_name":   a.campaign_name,
        "spend":           a.spend,
        "results":         a.results,
        "cost_per_result": a.cost_per_result,
        "roas":            a.roas,
        "ctr":             a.ctr,
        "frequency":       a.frequency,
    } for a in all_ads[:5]]
    session["date_range"] = date_range
    session["summary"] = {
        "total_spend":         summary.total_spend,
        "total_impressions":   summary.total_impressions,
        "total_reach":         summary.total_reach,
        "total_clicks":        summary.total_clicks,
        "total_link_clicks":   summary.total_link_clicks,
        "total_results":       summary.total_results,
        "avg_ctr":             summary.avg_ctr,
        "avg_cpc":             summary.avg_cpc,
        "avg_cpm":             summary.avg_cpm,
        "avg_roas":            summary.avg_roas,
        "avg_frequency":       summary.avg_frequency,
        "avg_cost_per_result": summary.avg_cost_per_result,
        "num_campaigns":       summary.num_campaigns,
        "num_ad_sets":         summary.num_ad_sets,
        "num_ads":             summary.num_ads,
        "campaign_type":       summary.campaign_type,
        "top_ad":              summary.top_ad,
        "top_ad_set":          summary.top_ad_set,
        "worst_ad":            summary.worst_ad,
        "worst_ad_set":        summary.worst_ad_set,
        "worst_ad_cpl":        summary.worst_ad_cpl,
        "has_click_data":      summary.has_click_data,
    }
    # Do NOT store full insights in session cookie — Claude output can be 4 KB+
    # which pushes the signed cookie over the 4 KB browser limit, silently dropping
    # the entire session on the next request. Insights are persisted in DB instead.

    # ── Persist to DB if a client is active ──────────────────────────────────
    client_id = session.get("client_id")
    if client_id and db.is_available() and not skip_db_save:
        try:
            filename = Path(session.get("data_source", "")).name
            upload_id = db.save_upload(
                client_id=client_id,
                filename=filename,
                date_from=date_range["from"] if date_range else None,
                date_to=date_range["to"] if date_range else None,
                total_spend=summary.total_spend,
                total_results=summary.total_results,
                avg_cpl=summary.avg_cost_per_result if summary.campaign_type != "purchases" else None,
                avg_roas=summary.avg_roas if summary.campaign_type == "purchases" else None,
                avg_ctr=summary.avg_ctr,
                avg_frequency=summary.avg_frequency,
                num_ads=summary.num_ads,
                campaign_type=summary.campaign_type,
                csv_content=csv_content,
            )
            session["upload_id"] = upload_id

            # Hook + format snapshots (use overrides so snapshots are correctly labelled)
            hook_perf = aggregate_hook_performance(all_ads, overrides=name_overrides)
            fmt_perf  = aggregate_format_performance(all_ads, overrides=name_overrides)
            db.save_hook_snapshots(client_id, upload_id, hook_perf)
            db.save_hook_snapshots(client_id, upload_id,
                                   [{**r, "hook_type": None} for r in fmt_perf])

            db.save_insights(client_id, upload_id, insights)
        except Exception as e:
            logger.warning("DB save failed: %s", e)

    # Detect unknown ads for tagging UI — alleen ads met voldoende spend
    sufficient_spend_ads = [a for a in all_ads if a.spend >= _LOW_SPEND_THRESHOLD]
    unknown_ads = get_unknown_ads(sufficient_spend_ads, overrides=name_overrides)
    tag_suggestions = suggest_ad_tags(unknown_ads) if unknown_ads else {}

    return {
        "summary":        summary,
        "all_ads_json":   all_ads_json,
        "ad_chart_data":  json.dumps(ad_chart_data),
        "insights_html":  _md_to_html(insights),
        "date_range":     date_range,
        "wow":            wow,
        "urgent_actions": urgent_actions,
        "unknown_ads":    unknown_ads,
        "tag_suggestions": tag_suggestions,
        "low_data_ads":      list(low_data_ads),
        "early_signal_ads":  list(early_signal_ads),
        "ad_delivery_map": ad_delivery_map,
    }


# ── Markdown → styled HTML ─────────────────────────────────────────────────────

_SECTION_STYLES = [
    (["sterke", "goed", "top", "winner", "schalen"],    "bi-check-circle-fill",          "#ecfdf5", "#059669"),
    (["verbeter", "under", "aandacht", "slecht", "stop", "kill"], "bi-exclamation-triangle-fill", "#fef2f2", "#dc2626"),
    (["aanbevel", "actie", "optimali", "tip", "hook", "creative"], "bi-lightbulb-fill",          "#fffbeb", "#d97706"),
    (["budget", "verdeling", "schaal", "reallocat"],    "bi-pie-chart-fill",              "#eff6ff", "#2563eb"),
]


def _section_style(title: str):
    t = title.lower()
    for keywords, icon, bg, color in _SECTION_STYLES:
        if any(k in t for k in keywords):
            return icon, bg, color
    return "bi-info-circle-fill", "#f8fafc", "#64748b"


def _render_body(text: str) -> str:
    lines = [l.rstrip() for l in text.strip().split("\n")]
    out, in_list = [], False
    for line in lines:
        if not line:
            continue
        if line.startswith(("- ", "* ", "• ")):
            if not in_list:
                out.append('<ul style="padding-left:1.1rem;margin:.25rem 0 0;">')
                in_list = True
            content = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", line[2:])
            out.append(f'<li style="margin-bottom:4px;line-height:1.5;">{content}</li>')
        else:
            if in_list:
                out.append("</ul>")
                in_list = False
            content = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", line)
            content = re.sub(r"\*(.+?)\*", r"<em>\1</em>", content)
            out.append(f'<p style="margin:0 0 4px;">{content}</p>')
    if in_list:
        out.append("</ul>")
    return "\n".join(out)


def _md_to_html(text: str) -> str:
    parts = re.split(r"^#{1,4}\s+(.+)$", text, flags=re.MULTILINE)
    html = []
    if parts[0].strip():
        html.append(f'<p style="font-size:.82rem;color:#374151;">{parts[0].strip()}</p>')
    for i in range(1, len(parts), 2):
        title   = parts[i].strip()
        content = parts[i + 1] if i + 1 < len(parts) else ""
        icon, bg, color = _section_style(title)
        body = _render_body(content)
        html.append(f"""
<div style="background:{bg};border-radius:9px;padding:.85rem 1rem;margin-bottom:.6rem;">
  <div style="font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:{color};margin-bottom:.5rem;">
    <i class="bi {icon}" style="margin-right:5px;"></i>{title}
  </div>
  <div style="font-size:.8rem;color:#1f2937;">{body}</div>
</div>""")
    return "\n".join(html)


def _session_to_summary(data: dict):
    from models.campaign import AnalysisSummary
    return AnalysisSummary(
        total_spend=data["total_spend"],
        total_impressions=data["total_impressions"],
        total_reach=data["total_reach"],
        total_clicks=data["total_clicks"],
        total_link_clicks=data["total_link_clicks"],
        total_results=data["total_results"],
        avg_ctr=data["avg_ctr"],
        avg_cpc=data["avg_cpc"],
        avg_cpm=data["avg_cpm"],
        avg_roas=data["avg_roas"],
        avg_frequency=data["avg_frequency"],
        avg_cost_per_result=data["avg_cost_per_result"],
        num_campaigns=data["num_campaigns"],
        num_ad_sets=data["num_ad_sets"],
        num_ads=data["num_ads"],
        campaign_type=data.get("campaign_type", "leads"),
        top_ad=data.get("top_ad"),
        top_ad_set=data.get("top_ad_set"),
        worst_ad=data.get("worst_ad"),
        worst_ad_set=data.get("worst_ad_set"),
        worst_ad_cpl=data.get("worst_ad_cpl"),
        has_click_data=data.get("has_click_data", True),
        campaigns=[],
    )


def _load_rows_from_session() -> list | None:
    source = session.get("data_source")
    if not source:
        return None
    if source == "demo":
        return load_dummy_data()
    # Merged uploads (data_source = "merged:1,2,3")
    if source.startswith("merged:"):
        upload_ids = [int(x) for x in source[7:].split(",") if x.strip().isdigit()]
        all_rows: list = []
        seen_keys: set = set()
        for uid in sorted(upload_ids):
            try:
                csv_text = db.get_upload_csv_content(uid)
                if csv_text:
                    for row in parse_csv_string(csv_text):
                        key = _dedup_key(row)
                        if key not in seen_keys:
                            seen_keys.add(key)
                            all_rows.append(row)
            except Exception:
                logger.warning("Merged row reload: upload %s failed", uid)
        return all_rows if all_rows else None
    # DB-stored single upload (data_source = "db:123")
    if source.startswith("db:"):
        try:
            uid = int(source[3:])
            csv_text = db.get_upload_csv_content(uid)
            if csv_text:
                return parse_csv_string(csv_text)
        except Exception:
            logger.warning("DB row reload failed for source: %s", source)
        return None
    # Local file path
    try:
        return parse_csv(Path(source))
    except Exception:
        logger.exception("Fout bij opnieuw laden CSV: %s", source)
        return None


def _classify_ads(all_ads, summary):
    """Returns (winners, losers, early_signals).
    Only ads with spend >= €50 get winner/loser classification.
    Ads with €25-€50 spend get the 'Vroeg signaal' label instead.
    """
    is_leads = summary.campaign_type != "purchases"
    reliable = [a for a in all_ads if a.spend >= _EARLY_SIGNAL_THRESHOLD]
    early    = [a for a in all_ads if _LOW_SPEND_THRESHOLD <= a.spend < _EARLY_SIGNAL_THRESHOLD]

    if is_leads:
        avg = summary.avg_cost_per_result
        winners = [a for a in reliable if a.results > 0 and a.cost_per_result > 0 and a.cost_per_result < avg * 0.85]
        losers  = [a for a in reliable if a.results == 0 and a.spend >= _EARLY_SIGNAL_THRESHOLD]
        if not winners:
            winners = sorted([a for a in reliable if a.results > 0], key=lambda x: x.cost_per_result)[:3]
    else:
        avg = summary.avg_roas
        winners = [a for a in reliable if a.roas > avg * 1.2 and a.roas > 0]
        losers  = [a for a in reliable if a.roas < avg * 0.5 and a.spend >= _EARLY_SIGNAL_THRESHOLD]
        if not winners:
            winners = sorted([a for a in reliable if a.roas > 0], key=lambda x: x.roas, reverse=True)[:3]
    return winners[:4], losers[:3], early[:6]


def _extract_patterns(winner_results: list) -> dict:
    from collections import Counter
    hooks   = [w["decoded"].get("hook_type", "unknown") for w in winner_results]
    formats = [w["decoded"].get("format", "unknown") for w in winner_results]
    drivers = [w["decoded"].get("psychological_driver", "") for w in winner_results if w["decoded"].get("psychological_driver")]
    return {
        "hook_counts":     dict(Counter(hooks)),
        "format_counts":   dict(Counter(formats)),
        "dominant_hook":   Counter(hooks).most_common(1)[0][0] if hooks else None,
        "dominant_format": Counter(formats).most_common(1)[0][0] if formats else None,
        "drivers": drivers,
        "untested_formats": [f for f in ["ugc", "carousel", "static", "testimonial"] if f not in formats],
    }


# ── Auth routes ────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if not _auth_enabled():
        session["username"] = "dev"
        return redirect(url_for("clients"))
    if session.get("username"):
        return redirect(url_for("clients"))
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        users = _get_users()
        if username in users and hmac.compare_digest(users[username], password):
            session.permanent = True
            session["username"] = username
            return redirect(url_for("clients"))
        return render_template("login.html", error="Gebruikersnaam of wachtwoord onjuist.")
    return render_template("login.html", error=None)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── Client routes ──────────────────────────────────────────────────────────────

@app.route("/debug/upload/<int:upload_id>")
@login_required
def debug_upload(upload_id):
    """Diagnostic: show what's stored for a specific upload."""
    if not db.is_available():
        return "<pre>DB niet beschikbaar</pre>"
    try:
        csv_text = db.get_upload_csv_content(upload_id)
        if not csv_text:
            return f"<pre>Upload {upload_id}: csv_content is NULL (geen CSV opgeslagen)</pre>"
        lines = csv_text.splitlines()
        preview = "\n".join(lines[:10])
        rows = parse_csv_string(csv_text)
        total_spend = sum(float(r.get("spend", 0) or 0) for r in rows)
        return (
            f"<pre style='font-family:monospace;padding:1rem;font-size:.85rem;'>"
            f"Upload ID  : {upload_id}\n"
            f"CSV regels : {len(lines)}\n"
            f"Parsed rows: {len(rows)}\n"
            f"Totaal spend: €{total_spend:.2f}\n\n"
            f"--- Eerste 10 regels CSV ---\n{html_escape(preview)}\n</pre>"
        )
    except Exception as e:
        return f"<pre>Fout: {html_escape(str(e))}</pre>"


@app.route("/debug/db")
@login_required
def debug_db():
    url_raw = os.getenv("DATABASE_URL", "")
    url_safe = ""
    if url_raw:
        # mask password for display
        import re as _re
        url_safe = _re.sub(r"://([^:]+):([^@]+)@", r"://\1:***@", url_raw)
    err = db.get_connection_error()
    available = db.is_available()
    return f"""<pre style="font-family:monospace;padding:2rem;font-size:.9rem;">
DB Diagnostiek
==============
DATABASE_URL ingesteld : {'JA' if url_raw else 'NEE'}
DATABASE_URL (masked)  : {url_safe or '(leeg)'}
Pool beschikbaar       : {'JA' if available else 'NEE'}
Laatste fout           : {err or '(geen)'}
psycopg2 versie        : {_psycopg2_version()}
</pre>"""


def _psycopg2_version() -> str:
    try:
        import psycopg2
        return psycopg2.__version__
    except ImportError:
        return "NIET GEINSTALLEERD"


def _make_test_png(width=100, height=100) -> bytes:
    """Generate a valid RGB PNG programmatically (no PIL needed)."""
    import struct, zlib
    def chunk(ctype, data):
        c = ctype + data
        return struct.pack(">I", len(data)) + c + struct.pack(">I", zlib.crc32(c) & 0xFFFFFFFF)
    sig = b"\x89PNG\r\n\x1a\n"
    ihdr = chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0))
    raw = b"".join(b"\x00" + bytes([255, 100, 50] * width) for _ in range(height))
    idat = chunk(b"IDAT", zlib.compress(raw))
    iend = chunk(b"IEND", b"")
    return sig + ihdr + idat + iend


@app.route("/debug/vision", methods=["GET", "POST"])
@login_required
def debug_vision():
    """Upload een echte afbeelding en zie wat de vision API exact teruggeeft."""
    from core.ai_client import _VISION_MODEL, has_api
    from core.static_analyzer import detect_hook_from_image

    if request.method == "GET":
        return """
        <html><body style="font-family:monospace;padding:2rem;">
        <h2>Vision API test — upload echte afbeelding</h2>
        <form method="POST" enctype="multipart/form-data">
            <input type="file" name="img" accept="image/*" required><br><br>
            <button type="submit">Test vision</button>
        </form>
        </body></html>
        """

    f = request.files.get("img")
    if not f:
        return "Geen bestand", 400

    import base64, traceback
    import anthropic as _anthropic

    ext = f.filename.rsplit(".", 1)[-1].lower()
    media_type = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp"}.get(ext, "image/jpeg")
    image_data = f.read()

    lines = [
        f"Bestand       : {f.filename}",
        f"Media type    : {media_type}",
        f"Grootte       : {len(image_data)} bytes",
        f"Model         : {_VISION_MODEL}",
        "",
        "--- detect_hook_from_image() ---",
    ]

    try:
        result = detect_hook_from_image(image_data, media_type)
        lines.append(f"hook_type     : {result.get('hook_type')}")
        lines.append(f"visual_summary: {result.get('visual_summary')}")
        lines.append(f"pain_point    : {result.get('pain_point')}")
        if "_error" in result:
            lines.append(f"_error        : {result['_error']}")
    except Exception as e:
        lines.append(f"FOUT: {e}")
        lines.append(traceback.format_exc())

    output = "\n".join(lines)
    return f"<pre style='font-family:monospace;padding:2rem;font-size:.9rem;'>{output}</pre>"


@app.route("/clients")
@login_required
def clients():
    try:
        client_list = db.get_clients() if db.is_available() else []
    except Exception as e:
        logger.error("DB get_clients failed: %s", e)
        flash(f"Database fout: {e}", "danger")
        client_list = []
    if not db.is_available():
        err = db.get_connection_error()
        flash(f"Database niet bereikbaar — {err} (ga naar /debug/db voor details)", "danger")
    return render_template("clients.html", clients=client_list)


@app.route("/clients/new", methods=["POST"])
@login_required
def client_new():
    name = request.form.get("name", "").strip()
    if not name:
        flash("Naam is verplicht.", "danger")
        return redirect(url_for("clients"))
    try:
        cpl  = float(request.form["cpl_benchmark"])  if request.form.get("cpl_benchmark")  else None
        roas = float(request.form["roas_benchmark"]) if request.form.get("roas_benchmark") else None
        client_id = db.create_client(
            name=name,
            industry=request.form.get("industry", ""),
            campaign_type=request.form.get("campaign_type", "leads"),
            cpl_benchmark=cpl,
            roas_benchmark=roas,
            notes=request.form.get("notes", ""),
            client_context=request.form.get("client_context", ""),
        )
        session["client_id"] = client_id
        flash(f"{name} aangemaakt.", "success")
        return redirect(url_for("client_profile", client_id=client_id))
    except Exception as e:
        flash(f"Fout: {e}", "danger")
        return redirect(url_for("clients"))


@app.route("/clients/<int:client_id>")
@login_required
def client_profile(client_id):
    try:
        client = db.get_client(client_id)
        if not client:
            flash("Klant niet gevonden.", "danger")
            return redirect(url_for("clients"))
        session["client_id"] = client_id
        session.pop("guest_mode", None)
        uploads      = db.get_uploads(client_id)
        shoot_briefs = db.get_shoot_briefs(client_id)
        hook_perf    = db.get_all_hook_performance(client_id)
    except Exception as e:
        logger.error("DB client_profile failed: %s", e)
        flash(f"Database fout: {e}", "danger")
        return redirect(url_for("clients"))

    # Bereken live hoeveel ads nog content missen vanuit de laatste upload
    pending_count = 0
    try:
        if uploads and db.is_available():
            csv_content = db.get_upload_csv_content(uploads[0]["id"])
            if csv_content:
                from core.csv_parser import parse_csv_string
                rows = filter_zero_spend(parse_csv_string(csv_content))
                missing = _get_new_ad_names(client_id, rows)
                pending_count = len(missing)
                if missing:
                    session[f"pending_new_ads_{client_id}"] = missing
                else:
                    session.pop(f"pending_new_ads_{client_id}", None)
    except Exception:
        pass

    meta_connection = None
    transcripts = []
    client_totals = {"total_spend": 0.0, "total_results": 0, "avg_cpl": None}
    try:
        if db.is_available():
            meta_connection = db.get_meta_connection(client_id)
            transcripts     = db.get_transcripts(client_id, limit=5)
            client_totals   = db.get_correct_totals(client_id)
    except Exception:
        pass

    return render_template("client_profile.html",
                           client=client, uploads=uploads,
                           shoot_briefs=shoot_briefs, hook_perf=hook_perf,
                           pending_count=pending_count,
                           meta_connection=meta_connection,
                           transcripts=transcripts,
                           client_totals=client_totals)


@app.route("/clients/<int:client_id>/edit", methods=["POST"])
@login_required
def client_edit(client_id):
    try:
        cpl  = float(request.form["cpl_benchmark"])  if request.form.get("cpl_benchmark")  else None
        roas = float(request.form["roas_benchmark"]) if request.form.get("roas_benchmark") else None
        db.update_client(
            client_id=client_id,
            name=request.form.get("name", ""),
            industry=request.form.get("industry", ""),
            campaign_type=request.form.get("campaign_type", "leads"),
            cpl_benchmark=cpl,
            roas_benchmark=roas,
            notes=request.form.get("notes", ""),
            client_context=request.form.get("client_context", ""),
        )
        flash("Klant bijgewerkt.", "success")
    except Exception as e:
        flash(f"Fout: {e}", "danger")
    return redirect(url_for("client_profile", client_id=client_id))


@app.route("/clients/<int:client_id>/delete")
@login_required
def client_delete(client_id):
    try:
        db.delete_client(client_id)
        if session.get("client_id") == client_id:
            session.pop("client_id", None)
        flash("Klant verwijderd.", "success")
    except Exception as e:
        flash(f"Fout bij verwijderen: {e}", "danger")
    return redirect(url_for("clients"))


@app.route("/clients/<int:client_id>/merge", methods=["POST"])
@login_required
def client_merge_uploads(client_id):
    upload_ids = request.form.getlist("upload_ids", type=int)
    if len(upload_ids) < 2:
        flash("Selecteer minimaal 2 uploads om samen te voegen.", "warning")
        return redirect(url_for("client_profile", client_id=client_id))

    session["client_id"] = client_id
    name_overrides = _load_name_overrides()

    all_rows: list = []
    seen_keys: set = set()
    loaded_count = 0

    for uid in sorted(upload_ids):  # oldest first → consistent dedup order
        try:
            csv_text = db.get_upload_csv_content(uid)
            if not csv_text:
                flash(f"Upload {uid} heeft geen opgeslagen CSV — sla op via de Laad-knop.", "warning")
                continue
            rows = parse_csv_string(csv_text)
            added = 0
            for row in rows:
                key = _dedup_key(row)
                if key not in seen_keys:
                    seen_keys.add(key)
                    all_rows.append(row)
                    added += 1
            if added > 0 or rows:
                loaded_count += 1
        except Exception as e:
            logger.warning("Merge: upload %s laden mislukt: %s", uid, e)

    if not all_rows:
        if loaded_count > 0:
            flash(
                f"{loaded_count} upload(s) gevonden maar bevatten geen bruikbare rijen. "
                "Controleer of de CSV-exports spend-data bevatten voor de gekozen periode.",
                "warning",
            )
        else:
            flash(
                "Geen opgeslagen CSV gevonden voor de geselecteerde uploads. "
                "Laad elke upload eerst afzonderlijk via de Laad-knop.",
                "danger",
            )
        return redirect(url_for("client_profile", client_id=client_id))

    # Run analysis without saving a new DB upload (temporarily unset client_id)
    session.pop("client_id", None)
    result = _process_df(all_rows, name_overrides=name_overrides)
    session["client_id"] = client_id
    session.pop("upload_id", None)  # no single upload_id for a merged analysis

    if "error" in result:
        flash(result["error"], "danger")
        return redirect(url_for("client_profile", client_id=client_id))

    session.permanent = True
    thresholds = session.get("thresholds", {"winner": 25, "mid": 50, "preset": "auto"})
    session["thresholds"] = thresholds
    session["data_source"] = f"merged:{','.join(str(i) for i in upload_ids)}"

    client = db.get_client(client_id)
    return render_template("index.html", result=result, demo=False,
                           thresholds=thresholds, active_client=client,
                           unknown_ads=result.get("unknown_ads", []),
                           tag_suggestions=result.get("tag_suggestions", {}),
                           merged_uploads=loaded_count)


@app.route("/clients/<int:client_id>/uploads/<int:upload_id>/delete", methods=["POST"])
@login_required
def upload_delete(client_id, upload_id):
    try:
        db.delete_upload(upload_id)
        session.pop(f"pending_new_ads_{client_id}", None)
        flash("Upload verwijderd.", "success")
    except Exception as e:
        flash(f"Fout bij verwijderen: {e}", "danger")
    return redirect(url_for("client_profile", client_id=client_id))


@app.route("/clients/<int:client_id>/briefs/<int:brief_id>/pdf", methods=["GET"])
@login_required
def brief_pdf(client_id, brief_id):
    client = db.get_client(client_id) if db.is_available() else None
    briefs = db.get_shoot_briefs(client_id, limit=50) if db.is_available() else []
    brief = next((b for b in briefs if b["id"] == brief_id), None)
    if not brief:
        flash("Shoot brief niet gevonden.", "warning")
        return redirect(url_for("client_profile", client_id=client_id))
    pdf_bytes = generate_shoot_brief_pdf(brief["brief_json"], client_name=client.get("name", "") if client else "")
    safe_name = (client.get("name") or "shoot_brief").replace(" ", "_").lower() if client else "shoot_brief"
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"shoot_brief_{safe_name}_{brief['created_at'].strftime('%Y%m%d')}.pdf",
    )


@app.route("/clients/<int:client_id>/briefs/<int:brief_id>/delete", methods=["POST"])
@login_required
def brief_delete(client_id, brief_id):
    try:
        db.delete_shoot_brief(brief_id)
        flash("Shoot brief verwijderd.", "success")
    except Exception as e:
        flash(f"Fout bij verwijderen: {e}", "danger")
    return redirect(url_for("client_profile", client_id=client_id))


@app.route("/clients/<int:client_id>/go/<path:destination>")
@login_required
def client_go(client_id, destination):
    """Load the latest upload for a client into session, then redirect to destination."""
    if not db.is_available():
        flash("Database niet beschikbaar.", "danger")
        return redirect(url_for("client_profile", client_id=client_id))

    uploads = db.get_uploads(client_id)
    if not uploads:
        flash("Geen uploads gevonden voor deze klant. Upload eerst een CSV.", "warning")
        return redirect(url_for("client_profile", client_id=client_id))

    latest = uploads[0]
    session["client_id"] = client_id

    rows = None
    try:
        csv_text = db.get_upload_csv_content(latest["id"])
        if csv_text:
            rows = parse_csv_string(csv_text)
            session["data_source"] = f"db:{latest['id']}"
    except Exception:
        pass

    if rows is None:
        file_path = UPLOAD_FOLDER / latest["filename"] if latest.get("filename") else None
        if file_path and file_path.exists():
            try:
                rows = parse_csv(file_path)
                session["data_source"] = str(file_path)
            except Exception:
                pass

    if rows is None:
        flash("CSV niet meer beschikbaar. Upload de export opnieuw.", "warning")
        return redirect(url_for("client_profile", client_id=client_id))

    name_overrides = _load_name_overrides()
    result = _process_df(rows,
                         campaign_type_override=latest.get("campaign_type", ""),
                         date_from=latest.get("date_from") or "",
                         date_to=latest.get("date_to") or "",
                         name_overrides=name_overrides,
                         skip_db_save=True)
    session["upload_id"] = latest["id"]

    if destination == "analyse":
        if "error" in result:
            flash(result["error"], "danger")
            return redirect(url_for("client_profile", client_id=client_id))
        thresholds = session.get("thresholds", {"winner": 25, "mid": 50, "preset": "auto"})
        _client_obj = None
        try:
            _client_obj = db.get_client(client_id)
        except Exception:
            pass
        return render_template("index.html", result=result, demo=False,
                               thresholds=thresholds, active_client=_client_obj,
                               unknown_ads=result.get("unknown_ads", []),
                               tag_suggestions=result.get("tag_suggestions", {}))

    dest_map = {
        "creative":     url_for("creative"),
        "hooks":        url_for("hooks"),
        "shoot-brief":  url_for("hooks", mode="brief"),
        "export/pdf":   url_for("export_pdf"),
    }
    return redirect(dest_map.get(destination, url_for("client_profile", client_id=client_id)))


@app.route("/clients/<int:client_id>/load/<int:upload_id>")
@login_required
def client_load_upload(client_id, upload_id):
    """Try to reload a historical CSV and run analysis again."""
    session["client_id"] = client_id
    uploads = db.get_uploads(client_id)
    upload  = next((u for u in uploads if u["id"] == upload_id), None)
    if not upload:
        flash("Upload niet gevonden.", "danger")
        return redirect(url_for("client_profile", client_id=client_id))

    # 1. Try to load CSV content from database
    rows = None
    try:
        csv_text = db.get_upload_csv_content(upload_id)
        if csv_text:
            rows = parse_csv_string(csv_text)
            session["data_source"] = f"db:{upload_id}"
    except Exception as e:
        logger.warning("DB csv_content load failed: %s", e)

    # 2. Fall back to disk if DB content not available
    if rows is None:
        file_path = UPLOAD_FOLDER / upload["filename"] if upload.get("filename") else None
        if file_path and file_path.exists():
            try:
                rows = parse_csv(file_path)
                session["data_source"] = str(file_path)
            except Exception as e:
                logger.warning("Disk CSV load failed: %s", e)

    if rows is None:
        flash("CSV niet meer beschikbaar. Upload de export opnieuw.", "warning")
        return redirect(url_for("client_profile", client_id=client_id))

    session.permanent = True
    name_overrides = _load_name_overrides()
    result = _process_df(rows,
                         campaign_type_override=upload.get("campaign_type", ""),
                         date_from=upload.get("date_from") or "",
                         date_to=upload.get("date_to") or "",
                         name_overrides=name_overrides,
                         skip_db_save=True)  # avoid creating duplicate upload records
    if "error" in result:
        flash(result["error"], "danger")
        return redirect(url_for("client_profile", client_id=client_id))
    # Keep the original upload_id in session so hooks/creative link to the right upload
    session["upload_id"] = upload_id
    thresholds = session.get("thresholds", {"winner": 25, "mid": 50, "preset": "auto"})
    client = db.get_client(client_id)
    return render_template("index.html", result=result, demo=False,
                           thresholds=thresholds, active_client=client,
                           unknown_ads=result.get("unknown_ads", []),
                           tag_suggestions=result.get("tag_suggestions", {}))


# ── Main analysis routes ───────────────────────────────────────────────────────

@app.route("/")
@login_required
def home():
    return redirect(url_for("clients"))


@app.route("/analyse", methods=["GET"])
@login_required
def index():
    client = None
    client_id = session.get("client_id")
    if client_id and db.is_available():
        try:
            client = db.get_client(client_id)
        except Exception:
            pass
    return render_template("index.html", result=None, thresholds=None, active_client=client,
                           unknown_ads=[], tag_suggestions={})


@app.route("/guest")
@login_required
def guest():
    session.pop("client_id", None)
    session["guest_mode"] = True
    return redirect(url_for("index"))


@app.route("/demo", methods=["GET"])
@login_required
def demo():
    rows = load_dummy_data()
    result = _process_df(rows)
    if "error" in result:
        flash(result["error"], "danger")
        return redirect(url_for("index"))
    thresholds = {"winner": 25, "mid": 50, "preset": "auto"}
    session.permanent = True
    session["data_source"] = "demo"
    session["thresholds"] = thresholds
    return render_template("index.html", result=result, demo=True,
                           thresholds=thresholds, active_client=None,
                           unknown_ads=result.get("unknown_ads", []),
                           tag_suggestions=result.get("tag_suggestions", {}))


@app.route("/upload", methods=["POST"])
@login_required
def upload():
    if "csv_file" not in request.files:
        flash("Geen bestand geselecteerd.", "warning")
        return redirect(url_for("index"))
    file = request.files["csv_file"]
    if file.filename == "":
        flash("Geen bestand geselecteerd.", "warning")
        return redirect(url_for("index"))
    if not file.filename.lower().endswith(".csv"):
        flash("Alleen CSV bestanden zijn toegestaan.", "danger")
        return redirect(url_for("index"))

    save_path = UPLOAD_FOLDER / secure_filename(file.filename)
    file.save(save_path)
    try:
        rows = parse_csv(save_path)
    except Exception as e:
        flash(f"Fout bij inlezen CSV: {e}", "danger")
        return redirect(url_for("index"))

    max_rows = int(os.getenv("MAX_CSV_ROWS", 10000))
    if len(rows) > max_rows:
        flash(f"CSV bevat {len(rows):,} rijen — maximum is {max_rows:,}.", "danger")
        return redirect(url_for("index"))

    # Read raw text for persistent storage in DB
    try:
        csv_text = save_path.read_text(encoding="utf-8-sig")
    except Exception:
        csv_text = None

    campaign_type_override = request.form.get("campaign_type_override", "")

    # Auto-threshold: load client to use CPL benchmark as default
    _upload_client = None
    _upload_client_id = session.get("client_id")
    if _upload_client_id and db.is_available():
        try:
            _upload_client = db.get_client(_upload_client_id)
        except Exception:
            pass
    thresholds = _compute_thresholds(request.form, client=_upload_client)

    name_overrides = _load_name_overrides()
    result = _process_df(rows, campaign_type_override=campaign_type_override,
                         csv_content=csv_text, name_overrides=name_overrides)
    if "error" in result:
        flash(result["error"], "danger")
        return redirect(url_for("index"))

    session.permanent = True
    # Als de upload in de DB is opgeslagen, gebruik db: als source en verwijder het lokale bestand.
    # Dit voorkomt een storage-lek en zorgt dat reanalyze geen nieuwe upload-record aanmaakt.
    _saved_upload_id = session.get("upload_id")
    if _saved_upload_id:
        session["data_source"] = f"db:{_saved_upload_id}"
        try:
            save_path.unlink(missing_ok=True)
        except Exception:
            pass
    else:
        session["data_source"] = str(save_path)
    session["thresholds"] = thresholds

    client = None
    client_id = session.get("client_id")
    if client_id and db.is_available():
        try:
            client = db.get_client(client_id)
        except Exception:
            pass

    # Detecteer nieuwe ads zonder creative content (alleen met spend > 0)
    rows_with_spend = filter_zero_spend(rows)
    _check_new_ads_after_upload(client_id, rows_with_spend)
    new_ads_count = len(session.get(f"pending_new_ads_{client_id}", []))
    if new_ads_count and client_id:
        flash(Markup(
            f"{new_ads_count} nieuwe advertentie(s) gevonden zonder script/copy. "
            f'<a href="{url_for("new_ads_content", client_id=client_id)}" class="alert-link">Voeg content toe →</a>'
        ), "info")

    return render_template("index.html", result=result, demo=False,
                           thresholds=thresholds, active_client=client,
                           unknown_ads=result.get("unknown_ads", []),
                           tag_suggestions=result.get("tag_suggestions", {}),
                           new_ads_count=new_ads_count)


@app.route("/tag-ads", methods=["POST"])
@login_required
def tag_ads():
    client_id = session.get("client_id")
    if not client_id or not db.is_available():
        return {"ok": False, "error": "Geen actieve klant"}, 400
    try:
        mappings = json.loads(request.form.get("mappings", "[]"))
        saved = db.save_ad_name_mappings(client_id, mappings)
        return {"ok": True, "saved": saved}
    except Exception as e:
        return {"ok": False, "error": str(e)}, 500


@app.route("/reanalyze", methods=["POST"])
@login_required
def reanalyze():
    rows = _load_rows_from_session()
    if rows is None:
        flash("Sessie verlopen. Upload je CSV opnieuw.", "warning")
        return redirect(url_for("index"))
    date_from = request.form.get("date_from", "").strip()
    date_to   = request.form.get("date_to",   "").strip()
    campaign_type_override = request.form.get("campaign_type_override", "")
    thresholds = _compute_thresholds(request.form)
    name_overrides = _load_name_overrides()
    # For merged or db sources, skip saving a new upload record (avoid duplicates)
    _source = session.get("data_source", "")
    skip_save = _source.startswith("merged:") or _source.startswith("db:")
    result = _process_df(rows, campaign_type_override=campaign_type_override,
                         date_from=date_from, date_to=date_to,
                         name_overrides=name_overrides,
                         skip_db_save=skip_save)
    if "error" in result:
        flash(result["error"], "danger")
        return redirect(url_for("index"))
    session.permanent = True
    session["thresholds"] = thresholds

    client = None
    client_id = session.get("client_id")
    if client_id and db.is_available():
        try:
            client = db.get_client(client_id)
        except Exception:
            pass

    return render_template("index.html", result=result, demo=False,
                           thresholds=thresholds, active_client=client,
                           unknown_ads=result.get("unknown_ads", []),
                           tag_suggestions=result.get("tag_suggestions", {}))


@app.route("/export/pdf", methods=["GET"])
@login_required
def export_pdf():
    summary_data = session.get("summary")
    if not summary_data:
        flash("Geen actieve analyse. Upload eerst een CSV.", "warning")
        return redirect(url_for("index"))
    summary   = _session_to_summary(summary_data)
    top_ads   = session.get("top_ads", [])
    date_range = session.get("date_range")

    # Insights are NOT stored in session (cookie size limit). Load from DB.
    insights_raw = "Geen inzichten beschikbaar."
    client_id = session.get("client_id")
    upload_id = session.get("upload_id")
    if client_id and db.is_available():
        try:
            hist = db.get_insights_history(client_id, limit=10)
            if hist:
                # Prefer insights that match the current upload; fall back to most recent
                match = next((h for h in hist if h.get("upload_id") == upload_id), None)
                insights_raw = (match or hist[0])["insights_text"] or insights_raw
        except Exception:
            pass
    if insights_raw == "Geen inzichten beschikbaar.":
        # Last-resort: regenerate from session summary (no API call, uses fallback)
        try:
            insights_raw = generate_insights(summary)
        except Exception:
            pass

    pdf_bytes = generate_pdf(summary, insights_raw, top_ads=top_ads, date_range=date_range)
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name="meta_ads_rapport.pdf",
    )


@app.route("/clients/<int:client_id>/export/shoot-brief", methods=["GET"])
@login_required
def export_shoot_brief_pdf(client_id):
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "warning")
        return redirect(url_for("hooks"))
    briefs = db.get_shoot_briefs(client_id, limit=1) if db.is_available() else []
    if not briefs:
        flash("Nog geen shoot brief beschikbaar voor deze klant. Genereer er eerst een.", "warning")
        return redirect(url_for("hooks", mode="brief"))
    scripts = briefs[0]["brief_json"]
    pdf_bytes = generate_shoot_brief_pdf(scripts, client_name=client.get("name", ""))
    safe_name = (client.get("name") or "shoot_brief").replace(" ", "_").lower()
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"shoot_brief_{safe_name}.pdf",
    )


@app.route("/creative", methods=["GET"])
@login_required
def creative():
    summary_data = session.get("summary")
    if not summary_data:
        flash("Geen actieve analyse. Upload eerst een CSV.", "warning")
        return redirect(url_for("index"))

    rows = _load_rows_from_session()
    if rows is None:
        flash("Het bestand kon niet worden geladen. Upload je CSV opnieuw.", "warning")
        return redirect(url_for("index"))

    campaigns = build_campaigns(rows)
    summary   = build_summary(rows, campaigns)
    all_ads   = sorted(get_all_ads(campaigns), key=lambda a: a.spend, reverse=True)
    winners, losers, early_signals = _classify_ads(all_ads, summary)

    winner_results = []
    for ad in winners:
        key = _cache_key(ad)
        cached = _cache_get(key)
        if cached:
            winner_results.append({"ad": ad, **cached})
        else:
            decoded = decode_winner(ad, summary)
            axes    = map_axes(decoded, ad.ad_name, all_ads=all_ads)
            testkit = generate_testkit(ad.ad_name, decoded, axes)
            entry   = {"decoded": decoded, "axes": axes, "testkit": testkit}
            _cache_set(key, entry)
            winner_results.append({"ad": ad, **entry})

    loser_results = []
    for ad in losers:
        key = _cache_key(ad) + "_loser"
        cached = _cache_get(key)
        if cached:
            loser_results.append({"ad": ad, "decoded": cached})
        else:
            decoded = decode_loser(ad, summary)
            _cache_set(key, decoded)
            loser_results.append({"ad": ad, "decoded": decoded})

    patterns   = _extract_patterns(winner_results)
    thresholds = session.get("thresholds", {"winner": 25, "mid": 50, "preset": "auto"})
    active_client = None
    if session.get("client_id") and db.is_available():
        try:
            active_client = db.get_client(session["client_id"])
        except Exception:
            pass
    return render_template(
        "creative.html",
        summary=summary,
        winners=winner_results,
        losers=loser_results,
        early_signals=early_signals,
        patterns=patterns,
        thresholds=thresholds,
        is_demo=session.get("data_source") == "demo",
        active_client=active_client,
        date_range=session.get("date_range"),
    )


@app.route("/hooks", methods=["GET"])
@login_required
def hooks():
    summary_data = session.get("summary")
    if not summary_data:
        flash("Geen actieve analyse. Upload eerst een CSV.", "warning")
        return redirect(url_for("index"))

    rows = _load_rows_from_session()
    if rows is None:
        flash("Het bestand kon niet worden geladen. Upload je CSV opnieuw.", "warning")
        return redirect(url_for("index"))

    campaigns = build_campaigns(rows)
    summary   = build_summary(rows, campaigns)
    all_ads   = sorted(get_all_ads(campaigns), key=lambda a: a.spend, reverse=True)

    name_overrides   = _load_name_overrides()
    hook_perf        = aggregate_hook_performance(all_ads, overrides=name_overrides)
    fmt_perf         = aggregate_format_performance(all_ads, overrides=name_overrides)
    combos           = get_winning_combinations(all_ads, overrides=name_overrides)
    untested_hooks   = get_untested_hooks(all_ads, overrides=name_overrides)
    untested_formats = get_untested_formats(all_ads, overrides=name_overrides)

    top_ad = next((a for a in all_ads if a.results > 0 and a.cost_per_result > 0), None)
    client_id = session.get("client_id")
    _client = None
    if client_id and db.is_available():
        try:
            _client = db.get_client(client_id)
        except Exception:
            pass
    # Enricheer shoot brief context met opgeslagen creative content
    _creative_ctx_for_brief = ""
    if client_id and db.is_available():
        try:
            _creatives_brief = db.get_ad_creatives(client_id)
            if _creatives_brief:
                from core.generation import _format_creative_context
                from core.hook_analyzer import detect_format
                _VIDEO_FORMATS = {"reels", "ugc", "testimonial", "story",
                                  "product_demo", "before_after", "problem_solve"}
                # Winnende video ads (reels e.d.) als primaire scriptbasis voor shoot brief
                winning_video_ads = [
                    a.ad_name for a in all_ads
                    if a.results > 0 and detect_format(a.ad_name) in _VIDEO_FORMATS
                ][:6]
                # Alle winnende ads als bredere context
                winning_ads_for_brief = [a.ad_name for a in all_ads[:8] if a.results > 0]
                _creative_ctx_for_brief = _format_creative_context(
                    _creatives_brief, winning_ads_for_brief,
                    video_ad_names=winning_video_ads,
                )
        except Exception:
            pass

    thresholds = session.get("thresholds", {"winner": 25, "mid": 50})

    shoot_briefs = []
    if client_id and db.is_available():
        try:
            shoot_briefs = db.get_shoot_briefs(client_id, limit=10)
        except Exception:
            pass

    return render_template(
        "hooks.html",
        summary=summary,
        hook_perf=hook_perf,
        fmt_perf=fmt_perf,
        combos=combos,
        untested_hooks=untested_hooks,
        untested_formats=untested_formats,
        is_demo=session.get("data_source") == "demo",
        active_client=_client,
        date_range=session.get("date_range"),
        t_win=thresholds.get("winner", 30),
        t_mid=thresholds.get("mid", 50),
        shoot_briefs=shoot_briefs,
    )


@app.route("/clients/<int:client_id>/generate-shoot-brief", methods=["POST"])
@login_required
def generate_shoot_brief_async(client_id):
    """Async endpoint: generate shoot brief and return rendered HTML partial."""
    rows = _load_rows_from_session()
    if rows is None:
        return "Geen data beschikbaar. Laad eerst een CSV.", 400

    campaigns = build_campaigns(rows)
    summary   = build_summary(rows, campaigns)
    all_ads   = sorted(get_all_ads(campaigns), key=lambda a: a.spend, reverse=True)
    top_ad    = next((a for a in all_ads if a.results > 0 and a.cost_per_result > 0), None)

    _client = None
    if db.is_available():
        try:
            _client = db.get_client(client_id)
        except Exception:
            pass

    _creative_ctx = ""
    if db.is_available():
        try:
            _creatives = db.get_ad_creatives(client_id)
            if _creatives:
                from core.generation import _format_creative_context
                from core.hook_analyzer import detect_format
                _VIDEO_FORMATS = {"reels", "ugc", "testimonial", "story",
                                  "product_demo", "before_after", "problem_solve"}
                winning_video_ads = [
                    a.ad_name for a in all_ads
                    if a.results > 0 and detect_format(a.ad_name) in _VIDEO_FORMATS
                ][:6]
                winning_ads = [a.ad_name for a in all_ads[:8] if a.results > 0]
                _creative_ctx = _format_creative_context(
                    _creatives, winning_ads, video_ad_names=winning_video_ads,
                )
        except Exception:
            pass

    _base_context = (_client.get("client_context") or "" if _client else "")
    _full_context = _base_context + ("===CREATIVE_CONTEXT===\n" + _creative_ctx if _creative_ctx else "")

    shoot_brief = generate_shoot_brief(
        summary, all_ads, top_ad=top_ad,
        client_name=_client["name"] if _client else "",
        client_context=_full_context,
    )

    upload_id = session.get("upload_id")
    if db.is_available():
        try:
            db.save_shoot_brief(client_id, upload_id, shoot_brief)
        except Exception as e:
            logger.warning("Shoot brief save failed: %s", e)

    return render_template("_shoot_brief_partial.html", shoot_brief=shoot_brief)


def _parse_static_image_upload(request_obj):
    """Validate and read uploaded image from a request. Returns (image_data, media_type) or raises ValueError."""
    if "image" not in request_obj.files:
        raise ValueError("Geen afbeelding meegestuurd")
    f = request_obj.files["image"]
    if not f.filename:
        raise ValueError("Geen bestand geselecteerd")
    allowed_types = {"image/jpeg", "image/png", "image/webp", "image/gif"}
    media_type = f.content_type or "image/jpeg"
    if media_type not in allowed_types:
        raise ValueError("Alleen JPEG, PNG, WebP en GIF zijn toegestaan")
    image_data = f.read()
    if len(image_data) > 5 * 1024 * 1024:
        raise ValueError("Afbeelding mag maximaal 5MB zijn")
    return image_data, media_type


def _build_existing_copies(client_id: int, hook_perf_db: list[dict] | None) -> list[dict]:
    """Build existing_copies list from ad_creatives + hook performance for a client."""
    if not db.is_available():
        return []
    try:
        creatives = db.get_ad_creatives(client_id)
    except Exception:
        return []

    cpl_by_hook: dict[str, float] = {}
    results_by_hook: dict[str, int] = {}
    if hook_perf_db:
        for row in hook_perf_db:
            ht = row.get("hook_type") or ""
            if ht and row.get("overall_cpl"):
                cpl_by_hook[ht] = float(row["overall_cpl"])
            if ht and row.get("total_results"):
                results_by_hook[ht] = int(row["total_results"])

    result = []
    for ad_naam, creative in creatives.items():
        for copy_key in ("ad_copy_1", "ad_copy_2", "ad_copy_3"):
            copy_text = creative.get(copy_key, "").strip()
            if not copy_text:
                continue
            from core.hook_analyzer import detect_hook
            hook_type = detect_hook(ad_naam)
            result.append({
                "ad_name": ad_naam,
                "copy": copy_text,
                "hook_type": hook_type,
                "cpl": cpl_by_hook.get(hook_type),
                "results": results_by_hook.get(hook_type),
            })
    return result


@app.route("/hooks/analyze-static", methods=["POST"])
@login_required
def analyze_static_image():
    """Receive a static ad image, analyse with Claude Vision, return 2 copy variants."""
    try:
        image_data, media_type = _parse_static_image_upload(request)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    hook_perf = None
    top_ads = None
    client_name = ""
    client_context = ""
    existing_copies = []

    rows = _load_rows_from_session()
    if rows:
        from core.analysis import build_campaigns, get_all_ads
        campaigns = build_campaigns(rows)
        all_ads = sorted(get_all_ads(campaigns), key=lambda a: a.spend, reverse=True)
        name_overrides = _load_name_overrides()
        hook_perf = aggregate_hook_performance(all_ads, overrides=name_overrides)
        top_ads = [a for a in all_ads if a.results > 0 and a.cost_per_result > 0][:5]

    client_id = session.get("client_id")
    if client_id and db.is_available():
        try:
            _client = db.get_client(client_id)
            if _client:
                client_name = _client.get("name", "")
                client_context = _client.get("client_context", "")
            hook_perf_db = db.get_all_hook_performance(client_id)
            existing_copies = _build_existing_copies(client_id, hook_perf_db)
            if not hook_perf:
                hook_perf = [
                    {"hook_type": r["hook_type"], "cpl": r.get("overall_cpl"),
                     "results": r.get("total_results"), "avg_ctr": r.get("avg_ctr")}
                    for r in hook_perf_db if r.get("hook_type")
                ]
        except Exception:
            pass

    from core.static_analyzer import analyze_static
    result = analyze_static(
        image_data, media_type,
        client_name=client_name,
        client_context=client_context,
        hook_perf=hook_perf,
        top_ads=top_ads,
        existing_copies=existing_copies,
    )
    return jsonify(result)


@app.route("/clients/<int:client_id>/analyze-static", methods=["POST"])
@login_required
def analyze_static_image_client(client_id: int):
    """Client profile static image analyser — uses DB data instead of session."""
    try:
        image_data, media_type = _parse_static_image_upload(request)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    if not db.is_available():
        return jsonify({"error": "Database niet beschikbaar"}), 503

    try:
        client = db.get_client(client_id)
    except Exception as e:
        return jsonify({"error": f"Klant niet gevonden: {e}"}), 404

    if not client:
        return jsonify({"error": "Klant niet gevonden"}), 404

    client_name = client.get("name", "")
    client_context = client.get("client_context", "")

    try:
        hook_perf_db = db.get_all_hook_performance(client_id)
    except Exception:
        hook_perf_db = []

    hook_perf = [
        {"hook_type": r["hook_type"], "cpl": r.get("overall_cpl"),
         "results": r.get("total_results"), "avg_ctr": r.get("avg_ctr")}
        for r in hook_perf_db if r.get("hook_type")
    ]
    existing_copies = _build_existing_copies(client_id, hook_perf_db)

    from core.static_analyzer import analyze_static
    result = analyze_static(
        image_data, media_type,
        client_name=client_name,
        client_context=client_context,
        hook_perf=hook_perf or None,
        existing_copies=existing_copies or None,
    )
    return jsonify(result)


@app.route("/clients/<int:client_id>/generate-scripts", methods=["POST"])
@login_required
def generate_scripts_client(client_id: int):
    """Generate 2 AI video script recommendations for a client using all available client data."""
    if not db.is_available():
        return jsonify({"error": "Database niet beschikbaar"}), 503

    try:
        client = db.get_client(client_id)
    except Exception as e:
        return jsonify({"error": f"Klant niet gevonden: {e}"}), 404

    if not client:
        return jsonify({"error": "Klant niet gevonden"}), 404

    client_name    = client.get("name", "")
    client_context = client.get("client_context", "")

    try:
        hook_perf_db = db.get_all_hook_performance(client_id)
    except Exception:
        hook_perf_db = []

    hook_perf = [
        {"hook_type": r["hook_type"], "cpl": r.get("overall_cpl"),
         "results": r.get("total_results"), "avg_ctr": r.get("avg_ctr")}
        for r in hook_perf_db if r.get("hook_type")
    ]

    # Build existing scripts from ad_creatives
    existing_scripts = []
    try:
        creatives = db.get_ad_creatives(client_id)
        cpl_by_hook: dict = {}
        results_by_hook: dict = {}
        for r in hook_perf_db:
            ht = r.get("hook_type") or ""
            if ht:
                if r.get("overall_cpl"):  cpl_by_hook[ht]     = float(r["overall_cpl"])
                if r.get("total_results"): results_by_hook[ht] = int(r["total_results"])

        from core.hook_analyzer import detect_hook
        for ad_naam, creative in creatives.items():
            script_text = creative.get("script", "").strip()
            if not script_text:
                continue
            hook_type = detect_hook(ad_naam)
            existing_scripts.append({
                "ad_name":     ad_naam,
                "script_text": script_text,
                "hook_type":   hook_type,
                "cpl":         cpl_by_hook.get(hook_type),
                "results":     results_by_hook.get(hook_type),
            })
    except Exception:
        pass

    from core.script_generator import generate_scripts
    result = generate_scripts(
        client_name=client_name,
        client_context=client_context,
        hook_perf=hook_perf or None,
        existing_scripts=existing_scripts or None,
    )
    return jsonify(result)


# ── Excel template downloads ───────────────────────────────────────────────────

@app.route("/templates/videos")
@login_required
def template_videos():
    """Download de lege videos Excel template."""
    try:
        data = generate_videos_template()
    except RuntimeError as e:
        flash(str(e), "danger")
        return redirect(request.referrer or url_for("clients"))
    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="template_videos.xlsx",
    )


@app.route("/templates/statics")
@login_required
def template_statics():
    """Download de lege statics Excel template."""
    try:
        data = generate_statics_template()
    except RuntimeError as e:
        flash(str(e), "danger")
        return redirect(request.referrer or url_for("clients"))
    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="template_statics.xlsx",
    )


# ── Bulk creative import (Excel templates) ────────────────────────────────────

@app.route("/clients/<int:client_id>/import/creatives", methods=["GET"])
@login_required
def import_creatives_page(client_id):
    """Toon de import-pagina voor bulk creative content."""
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))
    session["client_id"] = client_id
    creatives = db.get_ad_creatives(client_id) if db.is_available() else {}
    return render_template("import_creatives.html", client=client, creatives=creatives)


@app.route("/clients/<int:client_id>/import/creatives/export", methods=["GET"])
@login_required
def export_creatives_csv(client_id):
    """Download alle opgeslagen creative content als Excel-bestand."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))
    client_name = client["name"] if isinstance(client, dict) else client.name
    creatives = db.get_ad_creatives(client_id) if db.is_available() else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Creative Content"

    # Header styling
    header_fill   = PatternFill("solid", fgColor="0D1B2A")
    header_font   = Font(bold=True, color="FFFFFF", size=10)
    alt_fill      = PatternFill("solid", fgColor="F4F6FA")
    border_side   = Side(style="thin", color="E2E8F0")
    cell_border   = Border(bottom=Border(bottom=border_side).bottom)
    wrap_align    = Alignment(wrap_text=True, vertical="top")

    headers = ["Ad naam", "Script", "Headline 1", "Headline 2", "Headline 3",
               "Ad copy 1", "Ad copy 2", "Ad copy 3"]
    col_widths = [40, 65, 38, 38, 38, 55, 55, 55]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    for row_idx, (ad_naam, c) in enumerate(creatives.items(), start=2):
        values = [
            ad_naam,
            c.get("script") or "",
            c.get("headline") or "",
            c.get("headline_2") or "",
            c.get("headline_3") or "",
            c.get("ad_copy_1") or "",
            c.get("ad_copy_2") or "",
            c.get("ad_copy_3") or "",
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap_align
            if fill:
                cell.fill = fill
        # auto-hoogte: ruwweg 15pt per regel in script (max 120)
        script_lines = max(1, len(str(values[1]).split("\n")))
        ws.row_dimensions[row_idx].height = min(15 * script_lines, 120)

    # Freeze header row
    ws.freeze_panes = "A2"

    safe_name = re.sub(r"[^\w\-]", "_", client_name)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"creatives_{safe_name}.xlsx",
    )


@app.route("/clients/<int:client_id>/import/videos", methods=["POST"])
@login_required
def import_videos(client_id):
    """Verwerk een ingevulde videos Excel template."""
    return _handle_creative_import(client_id, "videos")


@app.route("/clients/<int:client_id>/import/statics", methods=["POST"])
@login_required
def import_statics(client_id):
    """Verwerk een ingevulde statics Excel template."""
    return _handle_creative_import(client_id, "statics")


def _handle_creative_import(client_id: int, template_type: str) -> object:
    file_key = f"{template_type}_file"
    if file_key not in request.files or request.files[file_key].filename == "":
        flash("Geen bestand geselecteerd.", "warning")
        return redirect(url_for("import_creatives_page", client_id=client_id))

    f = request.files[file_key]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        flash("Alleen Excel (.xlsx) bestanden zijn toegestaan.", "danger")
        return redirect(url_for("import_creatives_page", client_id=client_id))

    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))

    try:
        klantnaam, creatives = parse_template(f.stream, template_type)
    except Exception as e:
        flash(f"Fout bij inlezen Excel: {e}", "danger")
        return redirect(url_for("import_creatives_page", client_id=client_id))

    if not creatives:
        flash("Geen bruikbare rijen gevonden in het template. Zorg dat je Ad naam invult.", "warning")
        return redirect(url_for("import_creatives_page", client_id=client_id))

    # Optionele validatie: klantnaam uit B1 vergelijken met klant in DB
    if klantnaam and klantnaam.lower() != client["name"].lower():
        flash(
            f"Klantnaam in template ('{klantnaam}') komt niet overeen met klant '{client['name']}'. "
            "Import toch doorgezet — controleer of je het juiste bestand uploadt.",
            "warning",
        )

    try:
        saved = db.bulk_upsert_ad_creatives(client_id, creatives)
        type_label = "video scripts" if template_type == "videos" else "static headlines"
        flash(f"{saved} {type_label} opgeslagen voor {client['name']}.", "success")
    except Exception as e:
        flash(f"Database fout bij opslaan: {e}", "danger")

    return redirect(url_for("import_creatives_page", client_id=client_id))


# ── Bulk CSV upload (meerdere tegelijk) ───────────────────────────────────────

@app.route("/clients/<int:client_id>/import/csvs", methods=["POST"])
@login_required
def import_bulk_csvs(client_id):
    """Upload meerdere historische CSV's tegelijk — worden samengevoegd als één analyse."""
    files = request.files.getlist("csv_files")
    files = [f for f in files if f.filename and f.filename.lower().endswith(".csv")]

    if not files:
        flash("Selecteer minimaal één CSV bestand.", "warning")
        return redirect(url_for("client_profile", client_id=client_id))

    session["client_id"] = client_id
    name_overrides = _load_name_overrides()

    all_rows: list = []
    seen_keys: set = set()
    upload_ids_saved = []

    for f in files:
        try:
            raw_text = f.read().decode("utf-8-sig")
            rows = parse_csv_string(raw_text)
            rows = filter_zero_spend(rows)
            for row in rows:
                key = _dedup_key(row)
                if key not in seen_keys:
                    seen_keys.add(key)
                    all_rows.append(row)
        except Exception as e:
            flash(f"Fout bij inlezen '{f.filename}': {e}", "warning")

    if not all_rows:
        flash("Geen bruikbare data gevonden in de geüploade CSV's.", "danger")
        return redirect(url_for("client_profile", client_id=client_id))

    client = db.get_client(client_id) if db.is_available() else None
    thresholds = _compute_thresholds(request.form, client=client)

    # Reconstruct merged CSV for DB persistence so /creative and /hooks can reload data
    merged_csv_content = None
    if all_rows:
        _skip = {"_has_click_data"}
        _fields = [k for k in all_rows[0].keys() if k not in _skip]
        _buf = io.StringIO()
        _w = _csv_module.DictWriter(_buf, fieldnames=_fields, extrasaction="ignore")
        _w.writeheader()
        _w.writerows(all_rows)
        merged_csv_content = _buf.getvalue()

    result = _process_df(all_rows, name_overrides=name_overrides, csv_content=merged_csv_content)
    if "error" in result:
        flash(result["error"], "danger")
        return redirect(url_for("client_profile", client_id=client_id))

    session.permanent = True
    session["thresholds"] = thresholds
    # Use db: source so _load_rows_from_session can reload for /creative and /hooks
    _bulk_upload_id = session.get("upload_id")
    session["data_source"] = f"db:{_bulk_upload_id}" if _bulk_upload_id else "bulk_import"

    # Check op nieuwe ads zonder creative content
    new_ad_names = _get_new_ad_names(client_id, all_rows)
    if new_ad_names:
        session[f"pending_new_ads_{client_id}"] = new_ad_names
        flash(
            f"{len(new_ad_names)} nieuwe advertenties gevonden zonder script/copy. "
            "Je kunt ze direct invullen hieronder.",
            "info",
        )

    return render_template("index.html", result=result, demo=False,
                           thresholds=thresholds, active_client=client,
                           unknown_ads=result.get("unknown_ads", []),
                           tag_suggestions=result.get("tag_suggestions", {}),
                           merged_uploads=len(files))


# ── Ongoing tracking — nieuwe ads na upload ───────────────────────────────────

def _get_new_ad_names(client_id: int, rows: list[dict]) -> list[str]:
    """Geeft lijst van ad namen die in de rows zitten maar nog geen creative content hebben."""
    if not db.is_available():
        return []
    try:
        existing = db.get_ad_names_with_creatives(client_id)
        ad_names_in_csv = {r.get("ad_name", "") for r in rows if r.get("ad_name") and r.get("ad_name") != "Unknown"}
        return sorted(ad_names_in_csv - existing)
    except Exception:
        return []


@app.route("/clients/<int:client_id>/new-ads", methods=["GET"])
@login_required
def new_ads_content(client_id):
    """Toon formulier om content toe te voegen aan nieuwe ads na een CSV upload."""
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))
    pending = session.get(f"pending_new_ads_{client_id}", [])
    if not pending and db.is_available():
        # Sessie leeg (bijv. na herdeployment): herlaad vanuit laatste upload
        uploads = db.get_uploads(client_id)
        if uploads:
            last_upload_id = uploads[0]["id"]
            try:
                csv_content = db.get_upload_csv_content(last_upload_id)
                if csv_content:
                    from core.csv_parser import parse_csv_string
                    rows = filter_zero_spend(parse_csv_string(csv_content))
                    pending = _get_new_ad_names(client_id, rows)
                    if pending:
                        session[f"pending_new_ads_{client_id}"] = pending
            except Exception:
                pass
    if not pending:
        flash("Geen nieuwe advertenties gevonden die content nodig hebben.", "info")
        return redirect(url_for("client_profile", client_id=client_id))
    return render_template("new_ads_content.html", client=client, new_ads=pending)


@app.route("/clients/<int:client_id>/new-ads", methods=["POST"])
@login_required
def new_ads_content_save(client_id):
    """Sla ingevulde content op voor nieuwe ads."""
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))

    pending = session.get(f"pending_new_ads_{client_id}", [])
    saved_count = 0
    remaining = []

    for idx, ad_naam in enumerate(pending, start=1):
        script     = request.form.get(f"script_{idx}", "").strip()
        headline   = request.form.get(f"headline_{idx}", "").strip()
        headline_2 = request.form.get(f"headline2_{idx}", "").strip()
        headline_3 = request.form.get(f"headline3_{idx}", "").strip()
        copy1      = request.form.get(f"copy1_{idx}", "").strip()
        copy2      = request.form.get(f"copy2_{idx}", "").strip()
        copy3      = request.form.get(f"copy3_{idx}", "").strip()

        if any([script, headline, copy1]):
            try:
                db.upsert_ad_creative(client_id, ad_naam, script=script,
                                      headline=headline, headline_2=headline_2, headline_3=headline_3,
                                      ad_copy_1=copy1, ad_copy_2=copy2, ad_copy_3=copy3)
                saved_count += 1
            except Exception as e:
                logger.warning("Creative save failed for '%s': %s", ad_naam, e)
                remaining.append(ad_naam)
        else:
            remaining.append(ad_naam)

    if remaining:
        session[f"pending_new_ads_{client_id}"] = remaining
    else:
        session.pop(f"pending_new_ads_{client_id}", None)

    if saved_count:
        flash(f"Content opgeslagen voor {saved_count} advertentie(s).", "success")
    else:
        flash("Geen content ingevuld — overgeslagen.", "info")

    return redirect(url_for("client_profile", client_id=client_id))


@app.route("/clients/<int:client_id>/save-creative", methods=["POST"])
@login_required
def save_creative(client_id):
    """Sla creative content op voor één specifieke ad (AJAX of form POST)."""
    if not db.is_available():
        return {"ok": False, "error": "DB niet beschikbaar"}, 503
    try:
        ad_naam   = request.form.get("ad_naam", "").strip()
        if not ad_naam:
            return {"ok": False, "error": "Ad naam ontbreekt"}, 400
        db.upsert_ad_creative(
            client_id=client_id,
            ad_naam=ad_naam,
            script=request.form.get("script", "").strip(),
            headline=request.form.get("headline", "").strip(),
            headline_2=request.form.get("headline_2", "").strip(),
            headline_3=request.form.get("headline_3", "").strip(),
            ad_copy_1=request.form.get("ad_copy_1", "").strip(),
            ad_copy_2=request.form.get("ad_copy_2", "").strip(),
            ad_copy_3=request.form.get("ad_copy_3", "").strip(),
        )
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}, 500


# ── Creatives overzicht ────────────────────────────────────────────────────────

@app.route("/clients/<int:client_id>/creatives")
@login_required
def client_creatives(client_id):
    """Toon alle opgeslagen creative content van een klant."""
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))
    session["client_id"] = client_id
    creatives = db.get_ad_creatives(client_id) if db.is_available() else {}
    return render_template("import_creatives.html", client=client, creatives=creatives)


# ── Upload uitbreiden: detect nieuwe ads na CSV upload ────────────────────────

def _check_new_ads_after_upload(client_id: int | None, rows: list[dict]) -> None:
    """Detecteer nieuwe ads na een upload en sla ze op in de sessie."""
    if not client_id or not db.is_available():
        return
    new_ads = _get_new_ad_names(client_id, rows)
    if new_ads:
        session[f"pending_new_ads_{client_id}"] = new_ads
    else:
        session.pop(f"pending_new_ads_{client_id}", None)


# ── SRT / naam helpers ────────────────────────────────────────────────────────

def _parse_srt(srt_text: str) -> str:
    """Strip SRT timestamps and sequence numbers — return plain spoken text."""
    import re as _re
    lines = srt_text.splitlines()
    out = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.isdigit():
            continue
        if _re.match(r"^\d{2}:\d{2}:\d{2}[,\.]\d{3}\s*-->\s*\d{2}:\d{2}:\d{2}[,\.]\d{3}$", line):
            continue
        out.append(line)
    return " ".join(out).strip()


def _smart_version(client_id: int, format_type: str, hook_type: str, slug: str) -> int:
    """
    Versienummer: format + hook_type + minimaal 2/3 slug-woorden moeten matchen.
    Zelfde concept -> versie ophogen. Ander concept met zelfde hook -> v1.
    """
    import re as _re
    try:
        creatives = db.get_ad_creatives(client_id)
        prefix = format_type.lower() + "-" + hook_type.lower() + "-v"
        candidate_words = set(slug.split("-")) - {""}
        highest = 0
        for naam in creatives:
            normalized = _re.sub(r"[ ]*-[ ]*", "-", naam.lower())
            m = _re.match(_re.escape(prefix) + r"([0-9]+)-(.*)", normalized)
            if not m:
                continue
            version_num = int(m.group(1))
            existing_slug_words = set(m.group(2).split("-")) - {""}
            if len(candidate_words & existing_slug_words) >= 2:
                highest = max(highest, version_num)
        return highest + 1
    except Exception:
        return 1


_STOPWOORDEN = {
    "in", "de", "het", "een", "en", "op", "voor", "van", "dat", "is",
    "je", "ze", "dit", "maar", "dan", "ook", "er", "al", "zo", "nog",
    "wel", "niet", "met", "bij", "als", "om", "uit", "naar", "door",
    "over", "per", "tot", "aan", "wil", "jij", "hier", "eens", "die",
    "was", "zijn", "heb", "kan", "ik", "we", "te", "hij", "zij",
    "geen", "meer", "nu", "jouw", "mijn", "goed", "weer", "want",
    "want", "ook", "iets", "altijd", "nooit", "heel", "echt", "toch",
    "what", "this", "that", "with", "from", "text", "image", "the",
    "and", "for", "are", "you", "your", "not", "no", "yes", "its",
}


def _slug_words(text: str, n: int = 3) -> str:
    """Pak de n meest betekenisvolle woorden (stopwoorden gefilterd) als slug."""
    import re as _re
    words = _re.findall(r"[a-zA-Z0-9À-ɏ]+", text)
    meaningful = [w for w in words if w.lower() not in _STOPWOORDEN]
    return "-".join(w.lower() for w in meaningful[:n])


def _format_script(text: str) -> str:
    """Zet SRT-platte tekst om naar leesbare alinea's (Pure Python, geen AI)."""
    import re as _re
    sentences = _re.split(r"(?<=[.!?])\s+", text.strip())
    sentences = [s.strip() for s in sentences if s.strip()]
    if not sentences:
        return text
    paragraphs, chunk = [], []
    for s in sentences:
        chunk.append(s)
        if len(chunk) >= 3 or (len(chunk) >= 2 and len(s) < 50):
            paragraphs.append(" ".join(chunk))
            chunk = []
    if chunk:
        paragraphs.append(" ".join(chunk))
    return "\n\n".join(paragraphs)


def _get_winning_copies(client_id: int, max_n: int = 5) -> list[str]:
    """Laad de winnende ad_copy_1 velden gesorteerd op laagste CPL (heuristic: eerste paar)."""
    try:
        creatives = db.get_ad_creatives(client_id)
        copies = [v["ad_copy_1"] for v in creatives.values() if v.get("ad_copy_1")]
        return copies[:max_n]
    except Exception:
        return []


# ── Nieuwe advertentie flow ────────────────────────────────────────────────────

@app.route("/client/<int:client_id>/nieuwe-advertentie", methods=["GET"])
@login_required
def nieuwe_advertentie_get(client_id):
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))
    session["client_id"] = client_id
    test_mode = request.args.get("test") == "1"
    return render_template("nieuwe_advertentie.html", client=client, result=None, tab="video", test_mode=test_mode)


@app.route("/client/<int:client_id>/nieuwe-advertentie/video", methods=["POST"])
@login_required
def nieuwe_advertentie_video(client_id):
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))

    srt_file = request.files.get("srt_file")
    if not srt_file or not srt_file.filename.lower().endswith(".srt"):
        flash("Upload een geldig .srt bestand.", "danger")
        return render_template("nieuwe_advertentie.html", client=client, result=None, tab="video")

    try:
        srt_text = srt_file.read().decode("utf-8", errors="ignore")
    except Exception as e:
        flash(f"Bestand kon niet worden gelezen: {e}", "danger")
        return render_template("nieuwe_advertentie.html", client=client, result=None, tab="video")

    spoken_text = _parse_srt(srt_text)
    if not spoken_text:
        flash("Geen gesproken tekst gevonden in het SRT bestand.", "danger")
        return render_template("nieuwe_advertentie.html", client=client, result=None, tab="video")

    from core.ai_client import call_json, call_text, has_api
    result = {}

    # Stap 2 — Hook detectie
    hook_data = {"hook_type": "proof", "hook_explanation": "", "core_promise": "", "pain_point": ""}
    if has_api():
        hook_prompt = f"""Dit is de gesproken tekst van een Meta advertentie video.
Analyseer de openingszin en bepaal:
1. Hook type (kies uit: recognition, frustration, curiosity, proof, promise, confrontation, urgency, problem_solve, social_proof, educational)
2. Hook uitleg: waarom is dit dit hook type?
3. Kernbelofte in 1 zin
4. Aangesproken pijnpunt

Gesproken tekst:
{spoken_text[:1500]}

Geef terug als JSON: hook_type, hook_explanation, core_promise, pain_point"""
        hook_data = call_json(hook_prompt, max_tokens=600)
        if "_error" in hook_data or not hook_data.get("hook_type"):
            hook_data = {"hook_type": "proof", "hook_explanation": "Automatisch bepaald.", "core_promise": spoken_text[:80], "pain_point": ""}

    hook_type    = hook_data.get("hook_type", "proof").lower().replace(" ", "_")
    core_promise = hook_data.get("core_promise", "")

    # Stap 3 — Leesbaar script (pure Python, geen AI)
    script_formatted = _format_script(spoken_text)

    # Stap 4 — Naam genereren met slimme versiedetectie op basis van kernbelofte
    import re as _re_dbg
    _cp_words_raw = _re_dbg.findall(r"[a-zA-Z0-9À-ɏ]+", core_promise)
    _cp_filtered  = [w for w in _cp_words_raw if w.lower() not in _STOPWOORDEN]
    logger.info("DEBUG naam (video) | hook_type=%r | core_promise=%r | raw_words=%r | filtered=%r",
                hook_type, core_promise, _cp_words_raw[:8], _cp_filtered[:5])
    slug    = _slug_words(core_promise, 3) or _slug_words(spoken_text, 3) or "advertentie"
    versie  = _smart_version(client_id, "reels", hook_type, slug)
    ad_naam = f"reels-{hook_type}-v{versie}-{slug}"
    logger.info("DEBUG naam (video) | slug=%r | ad_naam=%r", slug, ad_naam)

    # Stap 5 — Copy genereren
    winning_copies = _get_winning_copies(client_id)
    client_context = client.get("client_context") or ""

    copy_data = {
        "ad_copy_1": "", "ad_copy_2": "", "ad_copy_3": "",
        "headline_1": "", "headline_2": "", "headline_3": "",
        "cta": "Plan een gratis kennismaking", "naam": ad_naam,
    }
    if has_api():
        prev_copies = "\n".join(f"  - {c}" for c in winning_copies) if winning_copies else "geen beschikbaar"
        copy_prompt = f"""Je bent een directe performance copywriter voor Meta Ads in Nederland.
Je taak: schrijf body copy die direct converteert.

Regels:
- Schrijf zoals mensen praten — niet zoals marketeers schrijven
- Gebruik de exacte woorden en zinnen uit het script/visual — niet parafraseren
- Geen clichés: niet "ontdek", "uniek", "effectief", "bewezen methode"
- Geen em-dashes
- Vanuit ik-perspectief of directe aanspraak (jij/je)
- Maximaal 3 zinnen per variant — elke zin telt
- Variant 3 is altijd maximaal 15 woorden — één punch
- De CTA is altijd laagdrempelig: passend bij het aanbod van deze klant — nooit "Koop nu" of "Schrijf je in"

Stijlreferentie — schrijf in dezelfde toon als dit script:
{spoken_text[:2000]}

Historisch beste copy van deze klant (gebruik dit als stijlreferentie, niet kopiëren):
{prev_copies}

Klantcontext:
{client_context[:600] if client_context else 'niet opgegeven'}

Geef terug als JSON:
{{
  "ad_copy_1": "maximaal 3 zinnen, bewezen aanpak gebaseerd op winnende stijl",
  "ad_copy_2": "maximaal 3 zinnen, andere emotionele invalshoek",
  "ad_copy_3": "maximaal 15 woorden, één directe punch",
  "headline_1": "maximaal 6 woorden, statement",
  "headline_2": "maximaal 6 woorden, andere invalshoek",
  "headline_3": "maximaal 8 woorden, vraagvorm",
  "cta": "laagdrempelige CTA passend bij het aanbod van deze klant"
}}"""
        copy_data = call_json(copy_prompt, max_tokens=1200)
        if "_error" in copy_data or not copy_data.get("ad_copy_1"):
            copy_data = {
                "ad_copy_1": "", "ad_copy_2": "", "ad_copy_3": "",
                "headline_1": "", "headline_2": "", "headline_3": "",
                "cta": "Plan een gratis kennismaking", "naam": ad_naam,
                "_error": copy_data.get("_error", "AI niet beschikbaar"),
            }
        copy_data["naam"] = ad_naam

    test_mode = request.form.get("test_mode") == "1"
    result = {
        "tab":              "video",
        "ad_naam":          ad_naam,
        "hook_type":        hook_type,
        "hook_explanation": hook_data.get("hook_explanation", ""),
        "core_promise":     core_promise,
        "pain_point":       hook_data.get("pain_point", ""),
        "script":           spoken_text,
        "script_formatted": script_formatted,
        "ad_copy_1":        copy_data.get("ad_copy_1", ""),
        "ad_copy_2":        copy_data.get("ad_copy_2", ""),
        "ad_copy_3":        copy_data.get("ad_copy_3", ""),
        "headline_1":       copy_data.get("headline_1", ""),
        "headline_2":       copy_data.get("headline_2", ""),
        "headline_3":       copy_data.get("headline_3", ""),
        "cta":              copy_data.get("cta", ""),
        "ai_error":         copy_data.get("_error"),
    }
    return render_template("nieuwe_advertentie.html", client=client, result=result, tab="video", test_mode=test_mode)


@app.route("/client/<int:client_id>/nieuwe-advertentie/static", methods=["POST"])
@login_required
def nieuwe_advertentie_static(client_id):
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))

    img_file = request.files.get("img_file")
    allowed_ext = {".jpg", ".jpeg", ".png", ".webp"}
    if not img_file or not any(img_file.filename.lower().endswith(ext) for ext in allowed_ext):
        flash("Upload een .jpg, .png of .webp afbeelding (max 5MB).", "danger")
        return render_template("nieuwe_advertentie.html", client=client, result=None, tab="static")

    img_data  = img_file.read()
    if len(img_data) > 5 * 1024 * 1024:
        flash("Afbeelding is groter dan 5MB.", "danger")
        return render_template("nieuwe_advertentie.html", client=client, result=None, tab="static")

    ext = img_file.filename.rsplit(".", 1)[-1].lower()
    media_type_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp"}
    media_type = media_type_map.get(ext, "image/jpeg")

    from core.ai_client import call_json, has_api
    from core.static_analyzer import detect_hook_from_image

    # Sla afbeelding op
    import uuid as _uuid
    img_filename = f"{client_id}_{_uuid.uuid4().hex[:8]}.{ext}"
    img_path = UPLOAD_FOLDER / img_filename
    try:
        img_path.write_bytes(img_data)
        afbeelding_pad = str(img_path)
    except Exception:
        afbeelding_pad = ""

    # Stap 1 — Vision: hook_type detectie (visual_summary is bonus, niet kritisch voor naam)
    hook_type      = "promise"
    visual_summary = ""
    pain_point     = ""
    client_context = client.get("client_context") or ""
    try:
        hook_data_raw  = detect_hook_from_image(img_data, media_type)
        logger.info("DEBUG static | detect_hook: %r", hook_data_raw)
        hook_type      = hook_data_raw.get("hook_type", "promise").lower().replace(" ", "_")
        visual_summary = hook_data_raw.get("visual_summary", "").strip()
        pain_point     = hook_data_raw.get("pain_point", "")
    except Exception as e:
        logger.warning("Vision analyse mislukt: %s", e)

    # Stap 2 — Copy genereren (VOOR naamgeneratie — headline wordt slug-fallback)
    winning_copies = _get_winning_copies(client_id)
    copy_data = {
        "ad_copy_1": "", "ad_copy_2": "", "ad_copy_3": "",
        "headline_1": "", "headline_2": "", "headline_3": "",
        "cta": "Plan een gratis kennismaking",
    }
    if has_api():
        prev_copies = "\n".join(f"  - {c}" for c in winning_copies) if winning_copies else "geen beschikbaar"
        copy_prompt = f"""Je bent een directe performance copywriter voor Meta Ads in Nederland.
Je taak: schrijf body copy die direct converteert.

Regels:
- Schrijf zoals mensen praten — niet zoals marketeers schrijven
- Gebruik de exacte woorden en zinnen uit het script/visual — niet parafraseren
- Geen clichés: niet "ontdek", "uniek", "effectief", "bewezen methode"
- Geen em-dashes
- Vanuit ik-perspectief of directe aanspraak (jij/je)
- Maximaal 3 zinnen per variant — elke zin telt
- Variant 3 is altijd maximaal 15 woorden — één punch
- De CTA is altijd laagdrempelig: passend bij het aanbod van deze klant — nooit "Koop nu" of "Schrijf je in"

Wat er op de afbeelding staat / visuele boodschap:
{visual_summary[:1000] if visual_summary else 'niet beschikbaar'}

Historisch beste copy van deze klant (gebruik dit als stijlreferentie, niet kopiëren):
{prev_copies}

Klantcontext:
{client_context[:600] if client_context else 'niet opgegeven'}

Geef terug als JSON:
{{
  "ad_copy_1": "maximaal 3 zinnen, bewezen aanpak gebaseerd op winnende stijl",
  "ad_copy_2": "maximaal 3 zinnen, andere emotionele invalshoek",
  "ad_copy_3": "maximaal 15 woorden, één directe punch",
  "headline_1": "maximaal 6 woorden, statement",
  "headline_2": "maximaal 6 woorden, andere invalshoek",
  "headline_3": "maximaal 8 woorden, vraagvorm",
  "cta": "laagdrempelige CTA passend bij het aanbod van deze klant"
}}"""
        copy_data = call_json(copy_prompt, max_tokens=1200)
        if "_error" in copy_data or not copy_data.get("ad_copy_1"):
            copy_data = {
                "ad_copy_1": "", "ad_copy_2": "", "ad_copy_3": "",
                "headline_1": "", "headline_2": "", "headline_3": "",
                "cta": "Plan een gratis kennismaking",
                "_error": copy_data.get("_error", "AI niet beschikbaar"),
            }

    # Stap 3 — Naam: visual_summary → headline_1 → hook_type (gegarandeerd betekenisvol)
    slug = _slug_words(visual_summary, 3)
    if not slug or slug == hook_type.replace("_", "-"):
        slug = _slug_words(copy_data.get("headline_1", ""), 3)
        logger.info("DEBUG static | slug uit headline: %r -> %r", copy_data.get("headline_1", ""), slug)
    if not slug:
        slug = hook_type.replace("_", "-")
    versie  = _smart_version(client_id, "static", hook_type, slug)
    ad_naam = f"static-{hook_type}-v{versie}-{slug}"
    logger.info("DEBUG static | ad_naam=%r", ad_naam)
    copy_data["naam"] = ad_naam

    test_mode = request.form.get("test_mode") == "1"
    result = {
        "tab":           "static",
        "ad_naam":       ad_naam,
        "hook_type":     hook_type,
        "visual_summary": visual_summary,
        "pain_point":    pain_point,
        "script":        visual_summary,
        "afbeelding_pad": afbeelding_pad,
        "ad_copy_1":     copy_data.get("ad_copy_1", ""),
        "ad_copy_2":     copy_data.get("ad_copy_2", ""),
        "ad_copy_3":     copy_data.get("ad_copy_3", ""),
        "headline_1":    copy_data.get("headline_1", ""),
        "headline_2":    copy_data.get("headline_2", ""),
        "headline_3":    copy_data.get("headline_3", ""),
        "cta":           copy_data.get("cta", ""),
        "ai_error":      copy_data.get("_error"),
    }
    return render_template("nieuwe_advertentie.html", client=client, result=result, tab="static", test_mode=test_mode)


@app.route("/client/<int:client_id>/nieuwe-advertentie/opslaan", methods=["POST"])
@login_required
def nieuwe_advertentie_opslaan(client_id):
    """Sla gegenereerde advertentie op in de database."""
    ad_naam      = request.form.get("ad_naam", "").strip()
    hook_type    = request.form.get("hook_type", "").strip()
    format_type  = request.form.get("format_type", "").strip()
    script       = request.form.get("script", "").strip()
    afbeelding   = request.form.get("afbeelding_pad", "").strip()
    ad_copy_1    = request.form.get("ad_copy_1", "").strip()
    ad_copy_2    = request.form.get("ad_copy_2", "").strip()
    ad_copy_3    = request.form.get("ad_copy_3", "").strip()
    headline_1   = request.form.get("headline_1", "").strip()
    headline_2   = request.form.get("headline_2", "").strip()
    headline_3   = request.form.get("headline_3", "").strip()
    cta          = request.form.get("cta", "").strip()
    test_mode    = request.form.get("test_mode") == "1"

    if not ad_naam:
        flash("Geen advertentienaam — opslaan geannuleerd.", "danger")
        return redirect(url_for("client_profile", client_id=client_id))

    # TEST MODUS — sla niets op, stuur terug met melding
    if test_mode:
        flash(f"[TEST] Advertentie '{ad_naam}' zou hier worden opgeslagen — geen DB-schrijf uitgevoerd.", "warning")
        return redirect(url_for("nieuwe_advertentie_get", client_id=client_id, test=1))

    if not db.is_available():
        flash("Database niet beschikbaar.", "danger")
        return redirect(url_for("client_profile", client_id=client_id))

    try:
        db.upsert_ad_creative(
            client_id=client_id,
            ad_naam=ad_naam,
            script=script,
            headline=headline_1,
            headline_2=headline_2,
            headline_3=headline_3,
            ad_copy_1=ad_copy_1,
            ad_copy_2=ad_copy_2,
            ad_copy_3=ad_copy_3,
            afbeelding_pad=afbeelding,
            hook_type=hook_type,
            format_type=format_type,
        )
        # Sla hook_type ook op in ad_name_mappings
        if hook_type:
            try:
                db.save_ad_name_mappings(client_id, [{"ad_name": ad_naam, "hook_type": hook_type, "format_type": format_type}])
            except Exception:
                pass
        flash(f"Advertentie '{ad_naam}' opgeslagen.", "success")
    except Exception as e:
        flash(f"Opslaan mislukt: {e}", "danger")

    return redirect(url_for("advertentie_geschiedenis", client_id=client_id))


# ── Advertentiegeschiedenis ────────────────────────────────────────────────────

@app.route("/client/<int:client_id>/advertentie-geschiedenis")
@login_required
def advertentie_geschiedenis(client_id):
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))
    session["client_id"] = client_id

    creatives = db.get_ad_creatives_list(client_id) if db.is_available() else []

    # Haal performance data op uit de meest recente upload voor CPL/classificatie
    ad_performance: dict[str, dict] = {}
    try:
        uploads = db.get_uploads(client_id)
        if uploads:
            csv_content = db.get_upload_csv_content(uploads[0]["id"])
            if csv_content:
                from core.csv_parser import parse_csv_string
                from core.analysis import build_campaigns, build_summary, get_all_ads
                rows = filter_zero_spend(parse_csv_string(csv_content))
                if rows:
                    camps = build_campaigns(rows)
                    summ  = build_summary(rows, camps)
                    all_ads = get_all_ads(camps)
                    winners, losers, early = _classify_ads(all_ads, summ)
                    winner_names  = {a.ad_name for a in winners}
                    loser_names   = {a.ad_name for a in losers}
                    early_names   = {a.ad_name for a in early}
                    for ad in all_ads:
                        cpl = ad.cost_per_result if ad.results > 0 else None
                        if ad.ad_name in winner_names:
                            label = "winner"
                        elif ad.ad_name in loser_names:
                            label = "loser"
                        elif ad.ad_name in early_names:
                            label = "vroeg_signaal"
                        else:
                            label = "middenmoter"
                        ad_performance[ad.ad_name] = {
                            "cpl": round(cpl, 2) if cpl else None,
                            "spend": round(ad.spend, 2),
                            "results": ad.results,
                            "label": label,
                        }
    except Exception:
        pass

    # Voeg meldingen toe aan elke creative
    for cr in creatives:
        naam  = cr.get("ad_naam", "")
        perf  = ad_performance.get(naam, {})
        has_script   = bool(cr.get("script"))
        has_copy     = bool(cr.get("ad_copy_1"))
        has_headline = bool(cr.get("headline"))
        if not has_script and not has_copy and not has_headline:
            cr["melding"] = "Nog geen content — voeg toe via Nieuwe advertentie"
            cr["melding_type"] = "leeg"
        elif has_script and not has_copy and not has_headline:
            cr["melding"] = "Script aanwezig — copy en headlines ontbreken nog"
            cr["melding_type"] = "deels"
        elif has_script and has_copy and not has_headline:
            cr["melding"] = "Headlines ontbreken nog"
            cr["melding_type"] = "bijna"
        else:
            cr["melding"] = ""
            cr["melding_type"] = "compleet"
        cr["cpl"]    = perf.get("cpl")
        cr["spend"]  = perf.get("spend")
        cr["results"] = perf.get("results")
        cr["label"]  = perf.get("label")

    return render_template("advertentie_geschiedenis.html",
                           client=client, creatives=creatives)


# ── Static brief ──────────────────────────────────────────────────────────────

@app.route("/client/<int:client_id>/static-brief")
@login_required
def static_brief(client_id):
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))
    session["client_id"] = client_id

    hook_perf     = []
    untested_hooks = []
    static_recs   = []

    try:
        uploads = db.get_uploads(client_id)
        if uploads:
            csv_content = db.get_upload_csv_content(uploads[0]["id"])
            if csv_content:
                from core.csv_parser import parse_csv_string
                from core.analysis import build_campaigns, build_summary, get_all_ads
                from core.hook_analyzer import aggregate_hook_performance, get_untested_hooks
                rows = filter_zero_spend(parse_csv_string(csv_content))
                if rows:
                    camps = build_campaigns(rows)
                    summ  = build_summary(rows, camps)
                    all_ads = get_all_ads(camps)
                    name_overrides = _load_name_overrides()
                    hook_perf = aggregate_hook_performance(all_ads, overrides=name_overrides)
                    # Welke hooks zijn NIET als static getest?
                    creatives = db.get_ad_creatives(client_id)
                    tested_static_hooks = {
                        v.get("hook_type") for v in creatives.values()
                        if v.get("hook_type") and v.get("format_type") == "static"
                    }
                    untested_hooks = [
                        r for r in hook_perf
                        if r["hook_type"] not in tested_static_hooks and r["hook_type"] != "unknown"
                    ]
                    # Aanbevolen statics: hoogst presterende hooks eerst
                    for row in hook_perf[:6]:
                        if row.get("results") and row["results"] > 0 and row["hook_type"] != "unknown":
                            static_recs.append({
                                "hook_type":    row["hook_type"],
                                "cpl":          row.get("cpl"),
                                "results":      row.get("results"),
                                "already_static": row["hook_type"] in tested_static_hooks,
                            })
    except Exception:
        pass

    return render_template("static_brief.html",
                           client=client,
                           hook_perf=hook_perf,
                           untested_hooks=untested_hooks,
                           static_recs=static_recs)


# ── Meta API integration ──────────────────────────────────────────────────────

try:
    from core.meta_api import (
        get_auth_url, exchange_code_for_token, get_ads,
        normalize_meta_data, refresh_token as _refresh_meta_token,
    )
    from core.ad_library import get_competitor_hooks, analyze_competitor_hooks, get_app_token
    from core.icp_updater import update_icp
    from core.action_planner import generate_action_plan
    from core.transcript_analyzer import analyze_and_save as analyze_transcript
    _META_AVAILABLE = True
except ImportError as _meta_import_err:
    logger.warning("Meta modules not available: %s", _meta_import_err)
    _META_AVAILABLE = False


def _encrypt_token(token: str) -> str:
    """Encrypt an access token using Fernet symmetric encryption."""
    key = os.getenv("TOKEN_ENCRYPTION_KEY", "")
    if not key:
        return token  # fall back to plaintext if no key configured
    try:
        from cryptography.fernet import Fernet
        f = Fernet(key.encode() if isinstance(key, str) else key)
        return f.encrypt(token.encode()).decode()
    except Exception as e:
        logger.warning("Token encryption failed, storing plaintext: %s", e)
        return token


def _decrypt_token(token: str) -> str:
    """Decrypt a Fernet-encrypted access token."""
    key = os.getenv("TOKEN_ENCRYPTION_KEY", "")
    if not key:
        return token
    try:
        from cryptography.fernet import Fernet
        f = Fernet(key.encode() if isinstance(key, str) else key)
        return f.decrypt(token.encode()).decode()
    except Exception:
        return token  # token was stored plaintext (no key at write time)


def _run_meta_sync(client_id: int, connection: dict,
                   date_from: str | None = None, date_to: str | None = None) -> dict:
    """
    Core sync logic shared by manual sync and cron sync.
    Returns a result dict with 'ok', 'upload_id', 'num_ads', 'error'.
    """
    from datetime import date, timedelta
    today = date.today()
    if not date_from:
        last_sync = connection.get("last_sync_at")
        if last_sync:
            if hasattr(last_sync, "date"):
                last_sync_date = last_sync.date()
            else:
                try:
                    last_sync_date = date.fromisoformat(str(last_sync)[:10])
                except ValueError:
                    last_sync_date = None
            if last_sync_date:
                date_from = (last_sync_date + timedelta(days=1)).isoformat()
        if not date_from:
            date_from = (today - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = today.isoformat()

    if date_from > date_to:
        logger.info("Sync skipped for client %s: already up to date (last=%s)", client_id, date_from)
        return {"ok": True, "upload_id": None, "num_ads": 0, "skipped": True}

    try:
        token    = _decrypt_token(connection["access_token"])
        acct_id  = connection["ad_account_id"]
        client   = db.get_client(client_id)

        raw_ads  = get_ads(token, acct_id, date_from, date_to)
        if not raw_ads:
            return {"ok": False, "error": "Meta API returned no ads for this period."}

        norm     = normalize_meta_data(raw_ads)

        total_spend   = sum(float(r.get("spend", 0) or 0) for r in norm)
        total_results = sum(int(r.get("results", 0) or 0) for r in norm)
        num_ads       = len(norm)
        avg_cpl  = round(total_spend / total_results, 2) if total_results > 0 else 0
        avg_roas = round(sum(float(r.get("roas", 0) or 0) for r in norm) / num_ads, 2) if num_ads else 0
        avg_ctr  = round(sum(float(r.get("ctr", 0) or 0) for r in norm) / num_ads, 2) if num_ads else 0
        avg_freq = round(sum(float(r.get("frequency", 0) or 0) for r in norm) / num_ads, 2) if num_ads else 0

        import csv as _csv_mod
        import io as _io
        buf = _io.StringIO()
        if norm:
            writer = _csv_mod.DictWriter(buf, fieldnames=list(norm[0].keys()))
            writer.writeheader()
            writer.writerows(norm)
        csv_content = buf.getvalue()

        upload_id = db.save_upload(
            client_id    = client_id,
            filename     = f"meta_sync_{date_from}_{date_to}.csv",
            date_from    = date_from,
            date_to      = date_to,
            total_spend  = total_spend,
            total_results= total_results,
            avg_cpl      = avg_cpl,
            avg_roas     = avg_roas,
            avg_ctr      = avg_ctr,
            avg_frequency= avg_freq,
            num_ads      = num_ads,
            campaign_type= client.get("campaign_type", "leads"),
            csv_content  = csv_content,
        )

        # Auto-tag + hook snapshots
        # aggregate_hook_performance expects Ad objects, so convert norm dicts first
        from core.hook_analyzer import aggregate_hook_performance
        from models.campaign import Ad as _Ad
        saved_mappings = db.get_ad_name_mappings(client_id)
        ad_objects = [
            _Ad(
                ad_id           = r.get("ad_id", ""),
                ad_name         = r.get("ad_name", ""),
                ad_set_name     = r.get("adset_name", ""),
                campaign_name   = r.get("campaign_name", ""),
                impressions     = int(r.get("impressions", 0) or 0),
                reach           = int(r.get("reach", 0) or 0),
                clicks          = int(r.get("clicks", 0) or 0),
                link_clicks     = int(r.get("link_clicks", 0) or 0),
                spend           = float(r.get("spend", 0) or 0),
                results         = int(r.get("results", 0) or 0),
                ctr             = float(r.get("ctr", 0) or 0),
                cpc             = float(r.get("cpc", 0) or 0),
                cpm             = float(r.get("cpm", 0) or 0),
                roas            = float(r.get("roas", 0) or 0),
                frequency       = float(r.get("frequency", 0) or 0),
                cost_per_result = float(r.get("cost_per_result", 0) or 0),
                delivery_status = r.get("status", ""),
            )
            for r in norm
        ]
        hook_perf = aggregate_hook_performance(ad_objects, overrides=saved_mappings)
        if hook_perf:
            db.save_hook_snapshots(client_id, upload_id, hook_perf)

        # Auto-update ICP
        try:
            update_icp(client_id, norm)
        except Exception as e:
            logger.warning("ICP update skipped: %s", e)

        db.update_last_sync(client_id)

        return {"ok": True, "upload_id": upload_id, "num_ads": num_ads}
    except Exception as e:
        logger.error("_run_meta_sync failed for client %s: %s", client_id, e)
        return {"ok": False, "error": str(e)}


@app.route("/client/<int:client_id>/connect-meta")
@login_required
def connect_meta(client_id):
    """Start the Meta OAuth flow for a client."""
    if not _META_AVAILABLE:
        flash("Meta API module niet geladen.", "danger")
        return redirect(url_for("client_profile", client_id=client_id))

    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))

    if not os.getenv("META_APP_ID"):
        flash("META_APP_ID omgevingsvariabele is niet ingesteld.", "danger")
        return redirect(url_for("client_profile", client_id=client_id))

    import secrets as _secrets
    oauth_nonce = _secrets.token_urlsafe(24)
    session["meta_oauth_nonce"] = oauth_nonce
    session["meta_oauth_pending_client_id"] = client_id
    auth_url = get_auth_url(state=oauth_nonce)
    return redirect(auth_url)


@app.route("/meta/callback")
@login_required
def meta_callback():
    """Receive the OAuth code from Meta, exchange for token, save to DB."""
    code       = request.args.get("code", "")
    state      = request.args.get("state", "")
    error      = request.args.get("error", "")

    if error:
        flash(f"Meta OAuth geweigerd: {request.args.get('error_description', error)}", "danger")
        return redirect(url_for("clients"))

    if not code:
        flash("Geen OAuth code ontvangen van Meta.", "danger")
        return redirect(url_for("clients"))

    # Validate state nonce to prevent CSRF
    expected_nonce = session.pop("meta_oauth_nonce", None)
    client_id = session.pop("meta_oauth_pending_client_id", None)
    if not expected_nonce or not hmac.compare_digest(state, expected_nonce) or not client_id:
        flash("Ongeldige OAuth state — mogelijk CSRF aanval. Probeer opnieuw.", "danger")
        return redirect(url_for("clients"))

    token_data = exchange_code_for_token(code)
    if "error" in token_data or "access_token" not in token_data:
        flash(f"Token uitwisseling mislukt: {token_data.get('error', 'onbekend')}", "danger")
        return redirect(url_for("client_profile", client_id=client_id))

    raw_token = token_data["access_token"]
    encrypted = _encrypt_token(raw_token)

    from datetime import datetime, timedelta
    expires_in = int(token_data.get("expires_in", 5184000))  # default 60 days
    expires_at = (datetime.utcnow() + timedelta(seconds=expires_in)).isoformat()

    # Fetch all ad accounts so the user can choose which one to link
    try:
        from core.meta_api import get_ad_accounts
        accounts = get_ad_accounts(raw_token)
    except Exception:
        accounts = []

    if not accounts:
        flash("Geen ad accounts gevonden voor dit Meta account. Controleer de app-rechten.", "danger")
        return redirect(url_for("client_profile", client_id=client_id))

    # Stash token + accounts in session — saved to DB only after account selection
    session["meta_oauth_token_enc"]  = encrypted
    session["meta_oauth_client_id"]  = client_id
    session["meta_oauth_expires_at"] = expires_at
    session["meta_oauth_accounts"]   = accounts

    return redirect(url_for("meta_select_account"))


@app.route("/meta/select-account", methods=["GET", "POST"])
@login_required
def meta_select_account():
    """Account selection step: show all ad accounts and let the user pick one."""
    client_id  = session.get("meta_oauth_client_id")
    accounts   = session.get("meta_oauth_accounts", [])
    encrypted  = session.get("meta_oauth_token_enc", "")
    expires_at = session.get("meta_oauth_expires_at", "")

    if not client_id or not accounts or not encrypted:
        flash("OAuth sessie verlopen. Start de koppeling opnieuw.", "warning")
        return redirect(url_for("clients"))

    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))

    if request.method == "POST":
        chosen_id = request.form.get("ad_account_id", "").strip()
        if not chosen_id:
            flash("Kies een ad account.", "warning")
            return render_template("meta_select_account.html",
                                   client=client, accounts=accounts)

        from datetime import datetime
        try:
            expires_dt = datetime.fromisoformat(expires_at) if expires_at else None
        except ValueError:
            expires_dt = None

        if db.is_available():
            db.save_meta_connection(client_id, chosen_id, encrypted, expires_dt)

        for key in ("meta_oauth_token_enc", "meta_oauth_client_id",
                    "meta_oauth_expires_at", "meta_oauth_accounts"):
            session.pop(key, None)

        flash("Meta Ads account succesvol gekoppeld!", "success")
        return redirect(url_for("meta_status", client_id=client_id))

    return render_template("meta_select_account.html", client=client, accounts=accounts)


@app.route("/client/<int:client_id>/meta-status")
@login_required
def meta_status(client_id):
    """Show Meta connection status and last sync time for a client."""
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))

    connection = db.get_meta_connection(client_id) if db.is_available() else None
    return render_template("meta_status.html", client=client, connection=connection)


@app.route("/client/<int:client_id>/sync-meta", methods=["POST"])
@login_required
def sync_meta(client_id):
    """Manually trigger a Meta data sync for a client."""
    if not _META_AVAILABLE:
        flash("Meta API module niet geladen.", "danger")
        return redirect(url_for("meta_status", client_id=client_id))

    if not db.is_available():
        flash("Database niet beschikbaar.", "danger")
        return redirect(url_for("meta_status", client_id=client_id))

    connection = db.get_meta_connection(client_id)
    if not connection:
        flash("Geen Meta verbinding gevonden. Koppel eerst je account.", "warning")
        return redirect(url_for("meta_status", client_id=client_id))

    date_from = request.form.get("date_from", "")
    date_to   = request.form.get("date_to", "")

    result = _run_meta_sync(client_id, connection, date_from or None, date_to or None)

    if result["ok"]:
        flash(f"Synchronisatie geslaagd: {result['num_ads']} advertenties opgehaald.", "success")
    else:
        flash(f"Synchronisatie mislukt: {result.get('error', 'onbekende fout')}", "danger")

    return redirect(url_for("meta_status", client_id=client_id))


@app.route("/client/<int:client_id>/meta-debug")
@login_required
def meta_debug(client_id):
    """
    Debug route: show raw Meta API responses for the stored token.
    Tests /me/adaccounts, /ads (simple), and /insights (what sync uses).
    Never used in production flows — diagnostic only.
    """
    import json as _json
    import requests as _requests
    from datetime import date, timedelta

    results = {}

    connection = db.get_meta_connection(client_id) if db.is_available() else None
    if not connection:
        return "<pre>Geen Meta verbinding gevonden voor deze klant.</pre>", 404

    token    = _decrypt_token(connection["access_token"])
    acct_id  = connection.get("ad_account_id", "act_498627026217281")
    base     = "https://graph.facebook.com/v19.0"
    date_to  = date.today().isoformat()
    date_from = (date.today() - timedelta(days=30)).isoformat()

    def _scrub(obj):
        """Recursively remove access_token values from the response before display."""
        if isinstance(obj, dict):
            return {
                k: ("[REDACTED]" if k == "access_token" else _scrub(v))
                for k, v in obj.items()
            }
        if isinstance(obj, list):
            return [_scrub(i) for i in obj]
        if isinstance(obj, str) and "access_token=" in obj:
            import re as _re
            return _re.sub(r"access_token=[^&\"]+", "access_token=[REDACTED]", obj)
        return obj

    def _call(label, url, params):
        try:
            r = _requests.get(url, params={**params, "access_token": token}, timeout=20)
            results[label] = {
                "status_code": r.status_code,
                "url":         url,
                "params":      {k: v for k, v in params.items() if k != "access_token"},
                "response":    _scrub(r.json()),
            }
        except Exception as e:
            results[label] = {"error": str(e), "url": url}

    # 1. Welke accounts zijn toegankelijk?
    _call(
        "1. GET /me/adaccounts",
        f"{base}/me/adaccounts",
        {"fields": "id,name,account_status,currency,timezone_name"},
    )

    # 2. Simpele ads lijst (zonder insights) — test basisrechten
    _call(
        "2. GET /ads (basis, geen insights)",
        f"{base}/{acct_id}/ads",
        {"fields": "id,name,status", "limit": 5},
    )

    # 3. Insights endpoint op ad-niveau — dit is wat de sync gebruikt
    _call(
        "3. GET /insights (level=ad, wat sync gebruikt)",
        f"{base}/{acct_id}/insights",
        {
            "level":      "ad",
            "fields":     "ad_id,ad_name,spend,impressions,clicks,actions,cost_per_action_type",
            "time_range": _json.dumps({"since": date_from, "until": date_to}),
            "limit":      5,
        },
    )

    # 4. Token info
    _call(
        "4. GET /me (token check)",
        f"{base}/me",
        {"fields": "id,name"},
    )

    pretty = _json.dumps(results, indent=2, ensure_ascii=False, default=str)
    html = (
        f"<!DOCTYPE html><html><head>"
        f"<meta charset='UTF-8'>"
        f"<title>Meta debug — client {client_id}</title>"
        f"<style>"
        f"body{{background:#0f172a;color:#e2e8f0;font-family:monospace;padding:2rem;max-width:1100px;margin:0 auto;}}"
        f"h1{{color:#ff5c2b;font-size:1.1rem;margin-bottom:.5rem;}}"
        f"p{{font-size:.8rem;color:#94a3b8;margin-bottom:1.5rem;}}"
        f"h2{{color:#94a3b8;font-size:.78rem;text-transform:uppercase;letter-spacing:.08em;"
        f"margin:1.5rem 0 .5rem;}}"
        f"pre{{background:#1e293b;padding:1.5rem;border-radius:12px;"
        f"overflow-x:auto;font-size:.8rem;line-height:1.6;white-space:pre-wrap;"
        f"border:1px solid rgba(255,255,255,.06);}}"
        f".warn{{color:#fbbf24;background:rgba(245,158,11,.1);border:1px solid rgba(245,158,11,.2);"
        f"padding:.6rem 1rem;border-radius:8px;font-size:.78rem;margin-bottom:1.25rem;}}"
        f".meta{{color:#64748b;font-size:.72rem;margin-bottom:1rem;}}"
        f"</style></head><body>"
        f"<h1>Meta API debug — client {client_id}</h1>"
        f"<p class='warn'>⚠ Access token is niet zichtbaar, maar responses bevatten account-IDs. "
        f"Deel deze output niet publiekelijk.</p>"
        f"<p class='meta'>Account: <strong style='color:#e2e8f0'>{acct_id}</strong> &nbsp;|&nbsp; "
        f"Periode: {date_from} → {date_to}</p>"
        f"<h2>Volledige response</h2>"
        f"<pre>{pretty}</pre>"
        f"</body></html>"
    )
    return html, 200


@app.route("/sync-all")
def sync_all():
    """
    Cron endpoint: sync all connected Meta accounts.
    Protected by X-Cron-Secret header matching CRON_SECRET env var.
    """
    cron_secret = os.getenv("CRON_SECRET", "")
    if not cron_secret:
        return jsonify({"error": "Unauthorized — CRON_SECRET not configured"}), 401
    if not hmac.compare_digest(request.headers.get("X-Cron-Secret", ""), cron_secret):
        return jsonify({"error": "Unauthorized"}), 401

    if not _META_AVAILABLE or not db.is_available():
        return jsonify({"error": "Meta module or DB not available"}), 503

    connections = db.get_all_meta_connections()
    results = []

    for conn in connections:
        client_id = conn["client_id"]
        res = _run_meta_sync(client_id, conn)
        results.append({
            "client_id":   client_id,
            "client_name": conn.get("client_name", ""),
            **res,
        })
        logger.info("Cron sync client %s: %s", client_id, res)

    ok_count  = sum(1 for r in results if r.get("ok"))
    err_count = len(results) - ok_count
    return jsonify({"synced": ok_count, "errors": err_count, "details": results})


@app.route("/client/<int:client_id>/action-plan")
@login_required
def action_plan(client_id):
    """Generate and display the weekly action plan for a client."""
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))

    plan_text = ""
    if db.is_available():
        plan_text = generate_action_plan(client_id)

    return render_template("action_plan.html", client=client, plan_text=plan_text)


@app.route("/client/<int:client_id>/competitor-analysis", methods=["GET", "POST"])
@login_required
def competitor_analysis(client_id):
    """Show competitor hook analysis from the Meta Ad Library."""
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))

    industry   = (client.get("industry") or "").strip()
    custom_kw  = request.values.get("keywords", "").strip()

    if custom_kw:
        keywords = [k.strip() for k in custom_kw.replace(",", "\n").splitlines() if k.strip()]
    else:
        keywords = [kw.strip() for kw in industry.split(",") if kw.strip()] or (
            [industry] if industry else []
        )

    # ── Token resolution (no separate env var needed) ──────────────────────
    token        = ""
    token_source = ""

    if _META_AVAILABLE:
        # 1. App access token: works with just META_APP_ID + META_APP_SECRET
        try:
            token = get_app_token()
            if token:
                token_source = "app"
        except Exception:
            pass

        # 2. Client's connected Meta user token
        if not token:
            try:
                if db.is_available():
                    conn = db.get_meta_connection(client_id)
                    if conn and conn.get("access_token"):
                        _t = _decrypt_token(conn["access_token"])
                        if _t:
                            token        = _t
                            token_source = "client"
            except Exception as _te:
                logger.warning("Could not get client Meta token: %s", _te)

    # 3. Explicit env var override
    if not token:
        token = os.getenv("META_AD_LIBRARY_TOKEN", "")
        if token:
            token_source = "env"

    library_disabled = not bool(token)

    competitor_ads = []
    analysis_text  = ""
    api_error      = ""

    if _META_AVAILABLE and keywords and token:
        competitor_ads, api_error = get_competitor_hooks(keywords, token=token)
        if competitor_ads:
            hook_perf    = db.get_all_hook_performance(client_id) if db.is_available() else []
            client_hooks = [h["hook_type"] for h in hook_perf if h.get("hook_type")]
            analysis_text = analyze_competitor_hooks(
                competitor_ads, client_hooks, client.get("name", "")
            )
        elif not api_error:
            api_error = f"Geen advertenties gevonden voor: {', '.join(keywords)}"

    return render_template(
        "competitor_analysis.html",
        client           = client,
        competitor_ads   = competitor_ads,
        analysis_text    = analysis_text,
        library_disabled = library_disabled,
        api_error        = api_error,
        token_source     = token_source,
        keywords_used    = keywords,
        custom_kw        = custom_kw,
    )


@app.route("/client/<int:client_id>/upload-transcript", methods=["POST"])
@login_required
def upload_transcript(client_id):
    """Accept a plain-text transcript upload, analyse it, and save to DB."""
    client = db.get_client(client_id) if db.is_available() else None
    if not client:
        flash("Klant niet gevonden.", "danger")
        return redirect(url_for("clients"))

    transcript_text = request.form.get("transcript_text", "").strip()
    if not transcript_text and "transcript_file" in request.files:
        f = request.files["transcript_file"]
        if f and f.filename:
            try:
                transcript_text = f.read().decode("utf-8", errors="replace").strip()
            except Exception as e:
                flash(f"Bestand kon niet worden gelezen: {e}", "danger")
                return redirect(url_for("client_profile", client_id=client_id))

    if not transcript_text:
        flash("Geen transcript ontvangen.", "warning")
        return redirect(url_for("client_profile", client_id=client_id))

    result = analyze_transcript(client_id, transcript_text)

    if "error" in result:
        flash(f"Analyse mislukt: {result['error']}", "danger")
    else:
        n = len(result.get("exact_phrases", []))
        flash(f"Transcript geanalyseerd: {n} hook-openingen gevonden.", "success")

    return redirect(url_for("client_profile", client_id=client_id))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
