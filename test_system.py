"""
System test — draait zonder Flask, DB of API key.
Test alle core modules op echte scenarios.
"""
import sys, io, traceback
sys.path.insert(0, ".")

PASS = 0
FAIL = 0

def check(name, condition, detail=""):
    global PASS, FAIL
    if condition:
        print(f"  ✓  {name}")
        PASS += 1
    else:
        print(f"  ✗  {name}" + (f"  →  {detail}" if detail else ""))
        FAIL += 1

def section(title):
    print(f"\n{'═'*60}\n  {title}\n{'═'*60}")


# ══════════════════════════════════════════════════════════════
# 1. CSV PARSER
# ══════════════════════════════════════════════════════════════
section("1. CSV Parser")
from core.csv_parser import parse_csv_string, validate_csv, COLUMN_MAP

# 1a. Basis NL kolommen
nl_csv = """Naam campagne,Naam advertentieset,Naam advertentie,Besteed bedrag (EUR),Vertoningen,Resultaten,Resultaatindicator
Camp A,Set 1,Ad Alpha,50.00,1000,5,lead
Camp A,Set 1,Ad Beta,0.00,500,0,lead
Camp B,Set 2,Ad Gamma,30.50,800,3,lead
"""
rows = parse_csv_string(nl_csv)
check("NL CSV: 3 rijen geparsed", len(rows) == 3)
check("NL CSV: spend correct", rows[0]["spend"] == 50.0)
check("NL CSV: campaign_name gemapped", rows[0]["campaign_name"] == "Camp A")
check("NL CSV: zero spend row aanwezig (niet gefilterd door parser)", rows[1]["spend"] == 0.0)

# 1b. EN kolommen
en_csv = """Campaign Name,Ad Set Name,Ad Name,Amount Spent (EUR),Impressions,Results,Result Indicator
Camp EN,Set EN,Ad EN,25.00,400,2,Lead
"""
rows_en = parse_csv_string(en_csv)
check("EN CSV: geparsed", len(rows_en) == 1)
check("EN CSV: campaign_name", rows_en[0]["campaign_name"] == "Camp EN")

# 1c. Summary-format (geen Day, geen IDs — Reporting Starts)
summary_csv = """Ad Name,Campaign Name,Ad Set Name,Amount Spent (EUR),Impressions,Results,Result Indicator,Reporting Starts
Ad X,Camp X,Set X,100.00,2000,10,lead,2025-01-01
Ad Y,Camp X,Set X,50.00,1000,5,lead,2025-01-01
"""
rows_s = parse_csv_string(summary_csv)
check("Summary CSV: geparsed", len(rows_s) == 2)
check("Summary CSV: day = Reporting Starts", rows_s[0]["day"] == "2025-01-01")

# 1d. Validate
ok, errs = validate_csv(rows)
check("Validate OK op geldige rijen", ok)
ok2, _ = validate_csv([])
check("Validate faalt op lege lijst", not ok2)


# ══════════════════════════════════════════════════════════════
# 2. ANALYSIS — merge_multi_conversion_rows + dedup
# ══════════════════════════════════════════════════════════════
section("2. Analysis — merge & dedup")
from core.analysis import (
    merge_multi_conversion_rows, filter_zero_spend,
    build_campaigns, build_summary, get_date_range,
)

# 2a. Meerdere conversion-rijen zelfde ad+dag → spend 1x, results optellen
dup_rows = [
    {"ad_id": "123", "ad_name": "Ad A", "campaign_id": "1", "campaign_name": "C",
     "ad_set_id": "10", "ad_set_name": "S", "day": "2025-01-01",
     "spend": 20.0, "results": 3, "impressions": 500, "clicks": 10,
     "link_clicks": 8, "reach": 400, "frequency": 1.2, "ctr": 2.0,
     "cpc": 2.0, "cpm": 40.0, "roas": 0.0, "cost_per_result": 6.67,
     "ctr_link": 0.0, "cpc_link": 0.0, "result_indicator": "lead",
     "ad_delivery": "", "_has_click_data": True},
    {"ad_id": "123", "ad_name": "Ad A", "campaign_id": "1", "campaign_name": "C",
     "ad_set_id": "10", "ad_set_name": "S", "day": "2025-01-01",
     "spend": 20.0, "results": 2, "impressions": 500, "clicks": 10,
     "link_clicks": 8, "reach": 400, "frequency": 1.2, "ctr": 2.0,
     "cpc": 2.0, "cpm": 40.0, "roas": 0.0, "cost_per_result": 10.0,
     "ctr_link": 0.0, "cpc_link": 0.0, "result_indicator": "custom_conversion",
     "ad_delivery": "", "_has_click_data": True},
]
merged = merge_multi_conversion_rows(dup_rows)
check("Merge: 2 dup-rijen → 1 rij", len(merged) == 1)
check("Merge: spend NIET verdubbeld", merged[0]["spend"] == 20.0)
check("Merge: results opgeteld (3+2=5)", merged[0]["results"] == 5.0)

# 2b. Summary-format: alle ad_id = "0" → dedup op naam
summary_rows = [
    {"ad_id": "0", "ad_name": "Ad Alpha", "campaign_id": "0", "campaign_name": "Unknown",
     "ad_set_id": "0", "ad_set_name": "Unknown", "day": "2025-01-01",
     "spend": 40.0, "results": 4, "impressions": 1000, "clicks": 20,
     "link_clicks": 15, "reach": 800, "frequency": 1.1, "ctr": 2.0,
     "cpc": 2.0, "cpm": 40.0, "roas": 0.0, "cost_per_result": 10.0,
     "ctr_link": 0.0, "cpc_link": 0.0, "result_indicator": "lead",
     "ad_delivery": "", "_has_click_data": True},
    {"ad_id": "0", "ad_name": "Ad Beta", "campaign_id": "0", "campaign_name": "Unknown",
     "ad_set_id": "0", "ad_set_name": "Unknown", "day": "2025-01-01",
     "spend": 20.0, "results": 2, "impressions": 500, "clicks": 10,
     "link_clicks": 8, "reach": 400, "frequency": 1.2, "ctr": 2.0,
     "cpc": 2.0, "cpm": 40.0, "roas": 0.0, "cost_per_result": 10.0,
     "ctr_link": 0.0, "cpc_link": 0.0, "result_indicator": "lead",
     "ad_delivery": "", "_has_click_data": True},
]
merged_s = merge_multi_conversion_rows(summary_rows)
check("Summary dedup: 2 verschillende ads blijven 2 rijen", len(merged_s) == 2)

# 2c. filter_zero_spend
mixed = [{"spend": 10.0}, {"spend": 0.0}, {"spend": 5.5}]
filtered = filter_zero_spend(mixed)
check("filter_zero_spend: verwijdert spend=0", len(filtered) == 2)
check("filter_zero_spend: houdt spend>0", all(r["spend"] > 0 for r in filtered))

# 2d. Volledige pipeline test
full_csv = """Campaign Name,Ad Set Name,Ad Name,Amount Spent (EUR),Impressions,Results,Result Indicator,Reach,Frequency
Camp A,Set 1,Ad One,80.00,2000,8,lead,1800,1.1
Camp A,Set 1,Ad Two,40.00,1000,2,lead,900,1.2
Camp A,Set 2,Ad Three,0.00,500,0,lead,450,1.0
Camp B,Set 3,Ad Four,60.00,1500,6,lead,1350,1.1
"""
rows_full = parse_csv_string(full_csv)
rows_full = filter_zero_spend(rows_full)
campaigns = build_campaigns(rows_full)
summary = build_summary(rows_full, campaigns)
check("Pipeline: spend=0 ads gefilterd (3 ads met spend)", summary.num_ads == 3)
check("Pipeline: totaal spend correct", summary.total_spend == 180.0)
check("Pipeline: totaal results correct", summary.total_results == 16)
check("Pipeline: 2 campaigns", summary.num_campaigns == 2)
check("Pipeline: campaign_type = leads", summary.campaign_type == "leads")
check("Pipeline: top_ad aanwezig", summary.top_ad is not None)


# ══════════════════════════════════════════════════════════════
# 3. HOOK ANALYZER — structured + fallback
# ══════════════════════════════════════════════════════════════
section("3. Hook Analyzer")
from core.hook_analyzer import (
    parse_ad_name, detect_hook, detect_format, detect_version,
    aggregate_hook_performance, get_unknown_ads,
)
from models.campaign import Ad

def make_ad(name, spend=50.0, results=5):
    return Ad(ad_id="1", ad_name=name, ad_set_name="Set", campaign_name="Camp",
              impressions=1000, reach=800, clicks=20, link_clicks=15,
              spend=spend, results=results, ctr=2.0, cpc=2.5, cpm=50.0,
              roas=0.0, frequency=1.2, cost_per_result=10.0)

# 3a. Structured format met underscore variants (de bug die we fixten)
cases = [
    ("reels - problem_solve - v1 - geen omkleden", "problem_solve", "reels", 1),
    ("reels - problem_solve - v2 - beschrijving",  "problem_solve", "reels", 2),
    ("static - social_proof - v1 - 500 klanten",   "social_proof",  "static", 1),
    ("ugc - frustration - v3 - moe van diëten",    "frustration",   "ugc", 3),
    ("testimonial - proof - v1 - klant review",    "proof",         "testimonial", 1),
    ("static - curiosity - v2 - wist je dat",      "curiosity",     "static", 2),
    ("reels - urgency - v1 - laatste kans",        "urgency",       "reels", 1),
    ("reels - educational - v1 - hoe werkt het",   "educational",   "reels", 1),
    ("static - confrontation - v1 - stop met",     "confrontation", "static", 1),
    ("static - promise - v1 - resultaat in 8w",    "promise",       "static", 1),
]
for name, exp_hook, exp_fmt, exp_ver in cases:
    parsed = parse_ad_name(name)
    check(f"Structured '{name[:40]}' → hook={exp_hook}",
          parsed["hook"] == exp_hook, f"got {parsed['hook']}")
    check(f"  format={exp_fmt}", parsed["format"] == exp_fmt, f"got {parsed['format']}")
    check(f"  version={exp_ver}", parsed["version"] == exp_ver, f"got {parsed['version']}")
    check(f"  structured=True", parsed["structured"] is True)

# 3b. Keyword fallback voor niet-gestructureerde namen
check("Fallback: 'Herken jij dit gevoel' → recognition",
      detect_hook("Herken jij dit gevoel als ondernemer?") == "recognition")
check("Fallback: 'Ziek van' → frustration",
      detect_hook("Ziek van diëten zonder resultaat?") == "frustration")
check("Fallback: '?' → curiosity",
      detect_hook("Weet jij al dit geheim?") == "curiosity")
check("Fallback: testimonial keyword → testimonial format",
      detect_format("Klantreview na 8 weken fit20") == "testimonial")

# 3c. Unknown ads (alleen ads met spend >= 30 en geen herkenbare hook)
ads = [
    make_ad("random naam zonder hook", spend=50.0),
    make_ad("reels - problem_solve - v1 - test", spend=50.0),
    make_ad("geen spend ad", spend=0.0),
    make_ad("te weinig data", spend=15.0),
]
unknown = get_unknown_ads([a for a in ads if a.spend >= 30])
unknown_names = [u["ad_name"] for u in unknown]
check("Unknown ads: onherkenbare naam gevonden", "random naam zonder hook" in unknown_names)
check("Unknown ads: gestructureerde ad niet in unknown", "reels - problem_solve - v1 - test" not in unknown_names)


# ══════════════════════════════════════════════════════════════
# 4. EXCEL TEMPLATES
# ══════════════════════════════════════════════════════════════
section("4. Excel Templates")
try:
    from core.excel_templates import generate_videos_template, generate_statics_template, parse_template

    # 4a. Genereer templates — moeten bytes teruggeven
    videos_bytes = generate_videos_template()
    statics_bytes = generate_statics_template()
    check("Videos template genereerd zonder fout", isinstance(videos_bytes, bytes) and len(videos_bytes) > 100)
    check("Statics template genereerd zonder fout", isinstance(statics_bytes, bytes) and len(statics_bytes) > 100)

    # 4b. Parse de gegenereerde template terug (geen data → lege lijst)
    klantnaam, creatives = parse_template(io.BytesIO(videos_bytes), "videos")
    check("Videos template parse: lege template → 0 creatives", len(creatives) == 0)

    # 4c. Schrijf data naar template en parse terug
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(videos_bytes))
    ws = wb.active
    ws["B1"] = "TestKlant"
    # Rij 5 = eerste echte datarij (rij 4 = voorbeeldrij)
    ws.cell(row=5, column=1, value="reels - problem_solve - v1 - test ad")
    ws.cell(row=5, column=2, value="Dit is het script voor de video")
    ws.cell(row=5, column=3, value="Ad copy 1 tekst")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    klantnaam2, creatives2 = parse_template(buf, "videos")
    check("Template parse: klantnaam correct", klantnaam2 == "TestKlant")
    check("Template parse: 1 creative gevonden", len(creatives2) == 1)
    check("Template parse: ad_naam correct", creatives2[0]["ad_naam"] == "reels - problem_solve - v1 - test ad")
    check("Template parse: script correct", creatives2[0]["script"] == "Dit is het script voor de video")
    check("Template parse: ad_copy_1 correct", creatives2[0]["ad_copy_1"] == "Ad copy 1 tekst")

    # 4d. Voorbeeldrij overgeslagen
    ws2 = load_workbook(io.BytesIO(videos_bytes)).active
    ws2["B1"] = "Klant"
    ws2.cell(row=4, column=1, value="Static - Proof - V1 - Klantresultaat fitness")  # voorbeeldrij
    ws2.cell(row=5, column=1, value="Mijn echte ad naam")
    buf2 = io.BytesIO()
    ws2.parent.save(buf2)
    buf2.seek(0)
    _, creatives3 = parse_template(buf2, "videos")
    check("Template parse: voorbeeldrij overgeslagen", len(creatives3) == 1)
    check("Template parse: alleen echte rij", creatives3[0]["ad_naam"] == "Mijn echte ad naam")

except ImportError:
    check("openpyxl beschikbaar", False, "openpyxl niet geïnstalleerd")
except Exception as e:
    check("Excel templates foutloos", False, str(e))
    traceback.print_exc()


# ══════════════════════════════════════════════════════════════
# 5. BULK CSV RECONSTRUCTIE (voor /creative + /hooks na bulk import)
# ══════════════════════════════════════════════════════════════
section("5. Bulk CSV reconstructie")
import csv as _csv_module

test_rows = [
    {"campaign_name": "C", "ad_set_name": "S", "ad_name": "Ad A",
     "spend": 50.0, "impressions": 1000, "results": 5,
     "result_indicator": "lead", "ad_delivery": "", "_has_click_data": True},
    {"campaign_name": "C", "ad_set_name": "S", "ad_name": "Ad B",
     "spend": 30.0, "impressions": 600, "results": 3,
     "result_indicator": "lead", "ad_delivery": "", "_has_click_data": True},
]
_skip = {"_has_click_data"}
_fields = [k for k in test_rows[0].keys() if k not in _skip]
_buf = io.StringIO()
_w = _csv_module.DictWriter(_buf, fieldnames=_fields, extrasaction="ignore")
_w.writeheader()
_w.writerows(test_rows)
csv_text = _buf.getvalue()

from core.csv_parser import parse_csv_string as pcs
reparsed = pcs(csv_text)
check("Reconstructie: CSV terug te lezen", len(reparsed) == 2)
check("Reconstructie: spend bewaard", reparsed[0]["spend"] == 50.0)
check("Reconstructie: ad_name bewaard", reparsed[0]["ad_name"] == "Ad A")
check("Reconstructie: geen _has_click_data kolom gelekt", "_has_click_data" not in _fields)


# ══════════════════════════════════════════════════════════════
# 6. DATUM + WoW VERGELIJKING
# ══════════════════════════════════════════════════════════════
section("6. Datum + WoW vergelijking")
from core.analysis import get_date_range, filter_rows_by_date, build_wow_comparison

date_rows = [
    {"day": "2025-01-01", "spend": 10.0, "results": 1, "impressions": 100, "clicks": 5},
    {"day": "2025-01-05", "spend": 20.0, "results": 2, "impressions": 200, "clicks": 10},
    {"day": "2025-01-10", "spend": 30.0, "results": 3, "impressions": 300, "clicks": 15},
    {"day": "2025-01-15", "spend": 40.0, "results": 4, "impressions": 400, "clicks": 20},
]
df, dt = get_date_range(date_rows)
check("Date range: from correct", df == "2025-01-01")
check("Date range: to correct", dt == "2025-01-15")

filtered_date = filter_rows_by_date(date_rows, "2025-01-05", "2025-01-10")
check("Date filter: 2 rijen binnen bereik", len(filtered_date) == 2)

wow = build_wow_comparison(date_rows)
check("WoW: geeft resultaat", wow is not None)
check("WoW: p1 en p2 aanwezig", "p1" in wow and "p2" in wow)
check("WoW: delta_spend aanwezig", "delta_spend" in wow)


# ══════════════════════════════════════════════════════════════
# 7. AD DELIVERY MAP
# ══════════════════════════════════════════════════════════════
section("7. Ad Delivery Map")
from core.analysis import build_ad_delivery_map

delivery_rows = [
    {"ad_name": "Ad Active",   "ad_delivery": "Active"},
    {"ad_name": "Ad Paused",   "ad_delivery": "Paused"},
    {"ad_name": "Ad Inactive", "ad_delivery": "Inactive"},
    {"ad_name": "Ad NoStatus", "ad_delivery": ""},
]
dmap = build_ad_delivery_map(delivery_rows)
check("Delivery map: active correct", dmap.get("Ad Active") == "active")
check("Delivery map: paused correct", dmap.get("Ad Paused") == "paused")
check("Delivery map: lege status niet opgenomen", "Ad NoStatus" not in dmap)


# ══════════════════════════════════════════════════════════════
# 8. GENERATION — prompt opbouw (zonder API)
# ══════════════════════════════════════════════════════════════
section("8. Generation — prompt opbouw")
from core.generation import _build_prompt, _format_creative_context

# Maak een minimale summary
from models.campaign import AnalysisSummary
dummy_summary = AnalysisSummary(
    total_spend=500.0, total_impressions=10000, total_reach=8000,
    total_clicks=200, total_link_clicks=150, total_results=20,
    avg_ctr=2.0, avg_cpc=2.5, avg_cpm=50.0, avg_roas=0.0,
    avg_frequency=1.5, avg_cost_per_result=25.0,
    num_campaigns=2, num_ad_sets=4, num_ads=8,
    campaign_type="leads",
    top_ad="Beste Ad", top_ad_set="Set 1",
    worst_ad="Slechtste Ad", worst_ad_set="Set 2",
    has_click_data=True, campaigns=[],
)
prompt = _build_prompt(dummy_summary)
check("Prompt: niet leeg", len(prompt) > 100)
check("Prompt: campaign type aanwezig", "leads" in prompt)
check("Prompt: spend aanwezig", "500" in prompt)

# Creative context
creatives = {
    "Beste Ad": {"script": "Dit is het script", "headline": "Headline tekst",
                 "ad_copy_1": "Copy 1", "ad_copy_2": "", "ad_copy_3": ""},
}
ctx = _format_creative_context(creatives, ["Beste Ad"])
check("Creative context: niet leeg", len(ctx) > 0)
check("Creative context: bevat WINNAAR label", "WINNAAR" in ctx)
check("Creative context: bevat script", "Dit is het script" in ctx)


# ══════════════════════════════════════════════════════════════
# EINDRESULTAAT
# ══════════════════════════════════════════════════════════════
print(f"\n{'═'*60}")
print(f"  RESULTAAT: {PASS} geslaagd  |  {FAIL} gefaald")
print(f"{'═'*60}\n")
sys.exit(0 if FAIL == 0 else 1)
