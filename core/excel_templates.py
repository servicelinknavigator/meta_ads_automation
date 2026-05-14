"""
Excel template generator en parser voor bulk creative import.
Genereert downloadbare templates en leest ingevulde templates in.
"""
from __future__ import annotations
import io
from typing import BinaryIO


def _make_workbook(columns: list[str], client_naam_label: str = "Klantnaam"):
    """Maak een openpyxl Workbook met instructierij en kolomheaders."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl is niet geïnstalleerd. Voeg 'openpyxl' toe aan requirements.txt")

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1C3557")
    meta_fill   = PatternFill("solid", fgColor="E8EEF7")
    meta_font   = Font(bold=True, color="1C3557")
    center      = Alignment(horizontal="center", vertical="center")

    # Rij 1: Klantnaam veld
    ws["A1"] = client_naam_label
    ws["A1"].font = meta_font
    ws["A1"].fill = meta_fill
    ws["A1"].alignment = center

    ws["B1"] = ""  # gebruiker vult hier de klantnaam in
    ws["B1"].font = Font(italic=True, color="888888")
    ws["B1"] = "← Vul hier de klantnaam in"
    ws.merge_cells("B1:E1")

    # Rij 2: lege scheiding
    ws.row_dimensions[2].height = 5

    # Rij 3: kolomheaders
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Rij 4+: voorbeeld (lichtgrijs)
    example_fill = PatternFill("solid", fgColor="F5F5F5")
    example_data = _example_row(columns)
    for col_idx, val in enumerate(example_data, start=1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        cell.fill = example_fill
        cell.font = Font(italic=True, color="999999")

    # Kolombreedte aanpassen
    col_widths = {"Ad naam": 40, "Script": 60, "Headline": 40,
                  "Ad copy 1": 50, "Ad copy 2": 50, "Ad copy 3": 50}
    for col_idx, col_name in enumerate(columns, start=1):
        width = col_widths.get(col_name, 30)
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = width

    ws.freeze_panes = "A4"
    return wb


def _example_row(columns: list[str]) -> list[str]:
    mapping = {
        "Ad naam":    "Static - Proof - V1 - Klantresultaat fitness",
        "Script":     "Heb jij ook het gevoel dat je al maanden traint zonder resultaat?...",
        "Headline":   "Van 0 naar resultaat in 8 weken",
        "Ad copy 1":  "Meer dan 500 klanten gingen je voor. Ontdek het programma dat écht werkt.",
        "Ad copy 2":  "Geen resultaat = geld terug. Zo zeker zijn wij van ons programma.",
        "Ad copy 3":  "",
    }
    return [mapping.get(c, "") for c in columns]


def generate_videos_template() -> bytes:
    """Genereer de videos Excel template als bytes."""
    columns = ["Ad naam", "Script", "Ad copy 1", "Ad copy 2", "Ad copy 3"]
    wb = _make_workbook(columns)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_statics_template() -> bytes:
    """Genereer de statics Excel template als bytes."""
    columns = ["Ad naam", "Headline", "Ad copy 1", "Ad copy 2", "Ad copy 3"]
    wb = _make_workbook(columns)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_template(file_obj: BinaryIO, template_type: str) -> tuple[str, list[dict]]:
    """
    Lees een ingevuld Excel template in.

    Returns:
        (klantnaam, list of creative dicts)
        klantnaam = "" als cel B1 leeg of standaard instructietekst bevat
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is niet geïnstalleerd")

    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    # Klantnaam uit B1
    raw_b1 = ws["B1"].value or ""
    klantnaam = str(raw_b1).strip()
    if "vul hier" in klantnaam.lower() or klantnaam == "":
        klantnaam = ""

    # Headers staan in rij 3
    headers = []
    for cell in ws[3]:
        val = str(cell.value or "").strip()
        headers.append(val)

    if not any(h for h in headers):
        return klantnaam, []

    # Data staat vanaf rij 4; rij 4 is voorbeeldrij als hij grijs/italic is — we lezen
    # alle rijen en filteren op Ad naam != leeg EN != voorbeeldwaarde
    creatives = []
    example_ad_naam = "Static - Proof - V1 - Klantresultaat fitness"

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(row):
            continue
        row_dict = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(headers)}

        ad_naam = row_dict.get("Ad naam", "").strip()
        if not ad_naam or ad_naam == example_ad_naam:
            continue

        creative: dict = {"ad_naam": ad_naam}
        if template_type == "videos":
            creative["script"]    = row_dict.get("Script", "")
            creative["ad_copy_1"] = row_dict.get("Ad copy 1", "")
            creative["ad_copy_2"] = row_dict.get("Ad copy 2", "")
            creative["ad_copy_3"] = row_dict.get("Ad copy 3", "")
        elif template_type == "statics":
            creative["headline"]  = row_dict.get("Headline", "")
            creative["ad_copy_1"] = row_dict.get("Ad copy 1", "")
            creative["ad_copy_2"] = row_dict.get("Ad copy 2", "")
            creative["ad_copy_3"] = row_dict.get("Ad copy 3", "")

        creatives.append(creative)

    return klantnaam, creatives
